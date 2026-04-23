"""
Agent Comptable IA - v5.0 SECURE
Traitement automatique de tickets de frais -> ecritures comptables Sage
Multi-provider : Claude -> OpenAI -> Ollama (fallback)
Securite : Auth, CSRF, Zero Data Retention, Anti-injection, Headers
"""

from dotenv import load_dotenv
load_dotenv()

import os
import io
import json
import base64
import re
import time
import email
import imaplib
import smtplib
import threading
import secrets
import hashlib
import hmac
import logging
from logging.handlers import RotatingFileHandler
from functools import wraps
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import unicodedata

import requests
from flask import (
    Flask, request, jsonify, render_template, send_file,
    session, redirect, url_for, abort, make_response,
    after_this_request
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.colors import red, black
from reportlab.lib.pagesizes import A4

app = Flask(__name__)


# ===================================================================
# LOGGING STRUCTURE
# ===================================================================

Path('logs').mkdir(exist_ok=True)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('enop')

handler = RotatingFileHandler('logs/enop.log', maxBytes=5*1024*1024, backupCount=3)
handler.setFormatter(logging.Formatter(
    '{"time": "%(asctime)s", "level": "%(levelname)s", "msg": "%(message)s"}'
))
logger.addHandler(handler)


# ===================================================================
# CONFIGURATION
# ===================================================================

# --- Securite ---
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FORCE_HTTPS', 'false').lower() == 'true'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)

# --- Identifiants login (via env, JAMAIS en dur) ---
APP_USERNAME = os.environ.get('APP_USERNAME', 'admin')
APP_PASSWORD_HASH = os.environ.get('APP_PASSWORD_HASH', '')
APP_PASSWORD_PLAIN = os.environ.get('APP_PASSWORD', 'changeme')

# --- API Keys ---
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', '')
GOOGLE_API_KEY = os.environ.get('GOOGLE_API_KEY', '')
GOOGLE_DOCAI_PROJECT_ID = os.environ.get('GOOGLE_DOCAI_PROJECT_ID', '')
GOOGLE_DOCAI_PROCESSOR_ID = os.environ.get('GOOGLE_DOCAI_PROCESSOR_ID', '')

# --- Modèle primaire vision (claude | gemini) ---
PRIMARY_LLM = os.getenv('PRIMARY_LLM', 'claude').lower().strip()
logger.info(f"  Primary LLM    : {PRIMARY_LLM}")

# --- Retry & Rate Limiting ---
MAX_RETRIES = 2
RETRY_BASE_DELAY = 1
RATE_LIMIT_DELAY = 0.5
RATE_LIMIT_429_WAIT = 15

# --- Brute-force protection ---
LOGIN_ATTEMPTS_FILE = Path('login_attempts.json')
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_DURATION = 300  # 5 minutes

# --- Rate limiting /api/process ---
PROCESS_RATE_LIMIT = {}

# --- Limite batch ---
MAX_PAGES_PER_BATCH = int(os.environ.get('MAX_PAGES_PER_BATCH', '50'))

# --- Webhook ---
WEBHOOK_TOKEN = os.environ.get('WEBHOOK_TOKEN', '')

# --- Dédup globale cross-page ---
DEDUP_ENABLED = os.environ.get('DEDUP_ENABLED', 'true').lower() == 'true'

# --- Dossiers ---
OUTPUT_FOLDER = Path('outputs')
OUTPUT_FOLDER.mkdir(exist_ok=True)
CACHE_FOLDER = Path('cache')
CACHE_FOLDER.mkdir(exist_ok=True)
CACHE_ENABLED = os.environ.get('CACHE_ENABLED', 'true').lower() == 'true'

# --- Auto-delete : supprimer les fichiers de plus de X minutes ---
FILE_RETENTION_MINUTES = int(os.environ.get('FILE_RETENTION_MINUTES', '10'))

# --- Email (optionnel) ---
EMAIL_ADDRESS = os.environ.get('EMAIL_ADDRESS', '')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD', '')
IMAP_SERVER = 'imap.gmail.com'
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
CHECK_INTERVAL = 30

# --- Prompts vision (externalises) ---
VISION_PROMPT = Path('prompts/vision_extraction.md').read_text(encoding='utf-8')
JUDGE_PROMPT = Path('prompts/vision_judge.md').read_text(encoding='utf-8')

# --- Addon JSON strict pour retry Gemini ---
JSON_STRICT_ADDON = """

DERNIER AVERTISSEMENT : ta reponse precedente etait cassee (JSON invalide ou tronque).

REGLES ABSOLUES CETTE FOIS :

1. Champ "raisonnement" : MAX 80 caracteres. PAS d apostrophe ('). PAS de guillemet.
   OK  : "7 tickets - 3 carburant 4 peages"
   KO  : "L'image contient 7 tickets..."

2. Tableau "tickets" : COMPLET. Si tu as annonce N dans nb_tickets_vus, mets N entrees.
   Ne tronque JAMAIS le tableau.

3. Aucun texte avant ou apres le JSON. Pas de ```json wrapper.

4. Point (.) comme separateur decimal. Virgules correctes.

Si tu doutes du raisonnement, mets "raisonnement": "" et concentre-toi sur
la completude du tableau tickets.
"""


# ===================================================================
# SECURITE : HELPERS
# ===================================================================

def hash_password(password):
    """Hash un mot de passe avec SHA-256 + salt"""
    salt = secrets.token_hex(16)
    h = hashlib.sha256(f"{salt}{password}".encode()).hexdigest()
    return f"{salt}:{h}"


def verify_password(password, stored_hash):
    """Verifie un mot de passe contre son hash"""
    if ':' not in stored_hash:
        return False
    salt, h = stored_hash.split(':', 1)
    return hmac.compare_digest(
        hashlib.sha256(f"{salt}{password}".encode()).hexdigest(),
        h
    )


def check_password(password):
    """Verifie le mot de passe (hash ou plain selon config)"""
    if APP_PASSWORD_HASH:
        return verify_password(password, APP_PASSWORD_HASH)
    return hmac.compare_digest(password, APP_PASSWORD_PLAIN)


def load_attempts():
    """Charge les tentatives de login depuis le fichier JSON"""
    if LOGIN_ATTEMPTS_FILE.exists():
        try:
            data = json.loads(LOGIN_ATTEMPTS_FILE.read_text(encoding='utf-8'))
            for ip, val in data.items():
                if val[1]:
                    data[ip][1] = datetime.fromisoformat(val[1])
            return data
        except Exception:
            return {}
    return {}


def save_attempts(attempts):
    """Sauvegarde les tentatives de login dans le fichier JSON"""
    data = {}
    for ip, val in attempts.items():
        data[ip] = [val[0], val[1].isoformat() if val[1] else None]
    LOGIN_ATTEMPTS_FILE.write_text(json.dumps(data, indent=2), encoding='utf-8')


def is_locked_out(ip):
    """Verifie si une IP est bloquee pour trop de tentatives"""
    attempts = load_attempts()
    if ip in attempts:
        count, lockout_time = attempts[ip]
        if lockout_time and datetime.now() < lockout_time:
            return True
        if lockout_time and datetime.now() >= lockout_time:
            del attempts[ip]
            save_attempts(attempts)
            return False
    return False


def record_failed_attempt(ip):
    """Enregistre une tentative de login echouee"""
    attempts = load_attempts()
    if ip not in attempts:
        attempts[ip] = [0, None]
    attempts[ip][0] += 1
    if attempts[ip][0] >= MAX_LOGIN_ATTEMPTS:
        attempts[ip][1] = datetime.now() + timedelta(seconds=LOCKOUT_DURATION)
    save_attempts(attempts)


def clear_attempts(ip):
    """Reset les tentatives apres un login reussi"""
    attempts = load_attempts()
    if ip in attempts:
        del attempts[ip]
        save_attempts(attempts)


def login_required(f):
    """Decorateur pour proteger les routes"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('authenticated'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Non authentifie'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def generate_csrf_token():
    """Genere un token CSRF"""
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']


def validate_csrf(token):
    """Valide le token CSRF"""
    return hmac.compare_digest(
        token or '',
        session.get('csrf_token', '')
    )


def sanitize_filename(filename):
    """Nettoie un nom de fichier contre les injections path traversal"""
    filename = os.path.basename(filename)
    filename = re.sub(r'[^\w\s\-\.]', '', filename)
    filename = filename.strip('. ')
    if not filename:
        filename = 'document.pdf'
    return filename


def cleanup_old_files():
    """Supprime les fichiers de sortie de plus de FILE_RETENTION_MINUTES"""
    try:
        cutoff = datetime.now() - timedelta(minutes=FILE_RETENTION_MINUTES)
        for f in OUTPUT_FOLDER.iterdir():
            if f.is_file():
                mtime = datetime.fromtimestamp(f.stat().st_mtime)
                if mtime < cutoff:
                    f.unlink()
                    logger.info(f"[Cleanup] Supprime {f.name}")
    except Exception as e:
        logger.error(f"[Cleanup] Erreur: {e}")


def schedule_cleanup():
    """Lance le nettoyage automatique toutes les 5 minutes"""
    while True:
        time.sleep(300)
        cleanup_old_files()


# ===================================================================
# SECURITE : MIDDLEWARE
# ===================================================================

@app.before_request
def security_checks():
    """Verifications de securite avant chaque requete"""
    # CSRF sur les POST (sauf login et webhook)
    if request.method == 'POST' and request.path not in ('/login', '/api/webhook'):
        if session.get('authenticated'):
            token = (request.form.get('csrf_token') or
                     request.headers.get('X-CSRF-Token') or
                     '')
            if not validate_csrf(token):
                abort(403)


@app.after_request
def security_headers(response):
    """Headers de securite sur toutes les reponses"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self'"
    )
    # Pas de cache sur les fichiers sensibles
    if request.path.startswith('/api/download'):
        response.headers['Cache-Control'] = 'no-store'
    return response


# ===================================================================
# UTILITAIRES PDF
# ===================================================================




def call_google_docai(pdf_bytes):
    """Extraction de texte via Google Document AI REST API."""
    import google.oauth2.service_account
    import google.auth.transport.requests

    if not GOOGLE_DOCAI_PROJECT_ID or not GOOGLE_DOCAI_PROCESSOR_ID:
        raise Exception("Google DocAI: GOOGLE_DOCAI_PROJECT_ID ou GOOGLE_DOCAI_PROCESSOR_ID non configure")

    credentials_path = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS', '')
    if not credentials_path or not os.path.exists(credentials_path):
        raise Exception("Google DocAI: fichier credentials introuvable")

    credentials = google.oauth2.service_account.Credentials.from_service_account_file(
        credentials_path,
        scopes=['https://www.googleapis.com/auth/cloud-platform']
    )
    auth_req = google.auth.transport.requests.Request()
    credentials.refresh(auth_req)
    access_token = credentials.token

    pdf_b64 = base64.b64encode(pdf_bytes).decode('utf-8')

    endpoint = (
        f"https://eu-documentai.googleapis.com/v1/projects/{GOOGLE_DOCAI_PROJECT_ID}"
        f"/locations/eu/processors/{GOOGLE_DOCAI_PROCESSOR_ID}:process"
    )

    response = requests.post(
        endpoint,
        headers={
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {access_token}'
        },
        json={
            'rawDocument': {
                'content': pdf_b64,
                'mimeType': 'application/pdf'
            }
        },
        timeout=60
    )

    if response.status_code == 200:
        raw_text = response.json().get('document', {}).get('text', '').strip()
        return {'full_text': raw_text}

    raise Exception(f"Google DocAI HTTP {response.status_code} - {response.text[:300]}")


def split_pdf_pages(pdf_bytes, filename):
    """Decoupe un PDF en pages individuelles"""
    reader = PdfReader(io.BytesIO(pdf_bytes))
    pages = []
    for i, page in enumerate(reader.pages):
        writer = PdfWriter()
        writer.add_page(page)
        output = io.BytesIO()
        writer.write(output)
        output.seek(0)
        page_name = f"{Path(filename).stem}_page{i+1}.pdf"
        pages.append({
            'filename': page_name,
            'bytes': output.read(),
            'original_filename': filename
        })
    return pages


def stamp_pdf_with_s(pdf_bytes):
    """Ajoute un S rouge sur le PDF"""
    reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()
    for page in reader.pages:
        packet = io.BytesIO()
        w = float(page.mediabox.width)
        h = float(page.mediabox.height)
        c = canvas.Canvas(packet, pagesize=(w, h))
        c.setFont("Helvetica-Bold", 60)
        c.setFillColor(red)
        c.setFillAlpha(0.7)
        c.drawString(w - 70, h - 70, "S")
        c.save()
        packet.seek(0)
        overlay = PdfReader(packet)
        page.merge_page(overlay.pages[0])
        writer.add_page(page)
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output.read()


def merge_pdfs(pdf_list):
    """Fusionne plusieurs PDFs en un seul"""
    writer = PdfWriter()
    for pdf_bytes in pdf_list:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        for page in reader.pages:
            writer.add_page(page)
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output.read()


# ===================================================================
# PROVIDERS IA
# ===================================================================

def render_page_as_png(pdf_bytes, dpi=300):
    """Rend la première page du PDF en PNG base64.
    Retourne (png_bytes, png_base64)."""
    import fitz
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    if len(doc) == 0:
        doc.close()
        raise ValueError("PDF vide")
    page = doc[0]
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat, alpha=False)
    png_bytes = pix.tobytes("png")
    doc.close()
    png_b64 = base64.b64encode(png_bytes).decode('utf-8')
    return png_bytes, png_b64


def call_gemini_vision(png_b64, extra_prompt=None):
    """Appel Gemini 3 Flash Vision sur une image PNG base64.
    extra_prompt : texte ajouté au system_instruction (pour retry avec contraintes renforcées).
    Retourne le texte brut de la réponse (à parser avec clean_json_response)."""
    if not GEMINI_API_KEY:
        raise Exception("Gemini: cle API non configuree")

    system_text = VISION_PROMPT + (extra_prompt or '')

    response = requests.post(
        f'https://generativelanguage.googleapis.com/v1beta/models/gemini-3-flash-preview:generateContent?key={GEMINI_API_KEY}',
        headers={'Content-Type': 'application/json'},
        json={
            'system_instruction': {'parts': [{'text': system_text}]},
            'contents': [{
                'parts': [
                    {'inline_data': {'mime_type': 'image/png', 'data': png_b64}},
                    {'text': 'Analyse cette image selon tes instructions.'}
                ]
            }],
            'generationConfig': {
                'temperature': 0.0,
                'maxOutputTokens': 8000,
                'response_mime_type': 'application/json',
            }
        },
        timeout=120
    )

    if response.status_code == 200:
        parts = response.json()['candidates'][0]['content']['parts']
        for part in parts:
            if part.get('thought'):
                continue
            if 'text' in part:
                return part['text']
        raise Exception("Gemini Vision: aucune part text dans la réponse")
    raise Exception(f"Gemini Vision HTTP {response.status_code} - {response.text[:300]}")


def compress_image_for_judge(png_bytes, max_size_mb=4.5):
    """Compresse une image PNG en JPEG si elle dépasse la limite Anthropic (5 MB base64).
    Pour rester sous 5 MB en base64, l'image binaire doit faire < 3.75 MB (base64 inflate ~33%).
    Retourne (image_bytes, media_type, base64_str)."""
    from io import BytesIO

    target_bytes = int(3.7 * 1024 * 1024)  # 3.7 MB binaire ≈ 4.9 MB base64

    if len(png_bytes) <= target_bytes:
        b64 = base64.b64encode(png_bytes).decode('utf-8')
        return png_bytes, 'image/png', b64

    # L'image est trop grosse : convertir en JPEG avec qualité adaptative
    try:
        from PIL import Image
    except ImportError:
        logger.warning("[Judge] Pillow non installé, impossible de compresser l'image")
        return png_bytes, 'image/png', base64.b64encode(png_bytes).decode('utf-8')

    img = Image.open(BytesIO(png_bytes))
    # Convertir en RGB si nécessaire (PNG peut être RGBA)
    if img.mode in ('RGBA', 'LA', 'P'):
        bg = Image.new('RGB', img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
        img = bg
    elif img.mode != 'RGB':
        img = img.convert('RGB')

    # Essayer plusieurs qualités JPEG
    for quality in (92, 85, 75, 65, 55):
        buf = BytesIO()
        img.save(buf, format='JPEG', quality=quality, optimize=True)
        jpg_bytes = buf.getvalue()
        if len(jpg_bytes) <= target_bytes:
            logger.info(
                f"[Judge] Image compressée PNG {len(png_bytes)//1024}KB → "
                f"JPEG q={quality} {len(jpg_bytes)//1024}KB"
            )
            return jpg_bytes, 'image/jpeg', base64.b64encode(jpg_bytes).decode('utf-8')

    # Si même q=55 ne suffit pas, redimensionner
    scale = 0.75
    while scale > 0.3:
        new_size = (int(img.width * scale), int(img.height * scale))
        small = img.resize(new_size, Image.LANCZOS)
        buf = BytesIO()
        small.save(buf, format='JPEG', quality=80, optimize=True)
        jpg_bytes = buf.getvalue()
        if len(jpg_bytes) <= target_bytes:
            logger.warning(
                f"[Judge] Image réduite à {int(scale*100)}% : "
                f"{len(jpg_bytes)//1024}KB JPEG q=80"
            )
            return jpg_bytes, 'image/jpeg', base64.b64encode(jpg_bytes).decode('utf-8')
        scale -= 0.15

    # Dernier recours : renvoyer tel quel
    logger.error(f"[Judge] Impossible de compresser sous {target_bytes//1024}KB")
    return png_bytes, 'image/png', base64.b64encode(png_bytes).decode('utf-8')


def call_claude_vision_primary(png_bytes, extra_prompt=None):
    """Claude Sonnet 4.6 Vision en mode PRIMAIRE (extraction tickets, thinking activé).
    Retourne le texte brut JSON (à parser avec clean_json_response).
    Même format de retour que call_gemini_vision() pour compatibilité pipeline."""
    if not ANTHROPIC_API_KEY:
        raise Exception("Anthropic: cle API non configuree")

    _, media_type, image_b64 = compress_image_for_judge(png_bytes)
    system_text = VISION_PROMPT + (extra_prompt or '')

    response = requests.post(
        'https://api.anthropic.com/v1/messages',
        headers={
            'Content-Type': 'application/json',
            'x-api-key': ANTHROPIC_API_KEY,
            'anthropic-version': '2023-06-01'
        },
        json={
            'model': 'claude-sonnet-4-6',
            'max_tokens': 8000,
            'thinking': {'type': 'enabled', 'budget_tokens': 4000},
            'system': system_text,
            'messages': [{
                'role': 'user',
                'content': [
                    {
                        'type': 'image',
                        'source': {
                            'type': 'base64',
                            'media_type': media_type,
                            'data': image_b64
                        }
                    },
                    {'type': 'text', 'text': 'Analyse cette image selon tes instructions.'}
                ]
            }]
        },
        timeout=180
    )

    if response.status_code == 200:
        content_blocks = response.json()['content']
        for block in reversed(content_blocks):
            if block.get('type') == 'text':
                return block['text']
        raise Exception("Claude Vision: aucun bloc text dans la réponse")

    error_msg = f"Claude Vision HTTP {response.status_code}"
    try:
        error_body = response.json()
        error_detail = error_body.get('error', {}).get('message', '')
        if error_detail:
            error_msg += f" - {error_detail}"
    except Exception:
        pass
    raise Exception(error_msg)


def call_claude_vision_judge(png_bytes, extraction_gemini):
    """Appel Claude Sonnet 4.6 Vision en mode judge.
    Compresse l'image en JPEG si elle dépasse la limite de 5 MB d'Anthropic."""
    if not ANTHROPIC_API_KEY:
        raise Exception("Anthropic: cle API non configuree")

    # Compression automatique si trop gros
    _, media_type, image_b64 = compress_image_for_judge(png_bytes)

    extraction_json = json.dumps(extraction_gemini, ensure_ascii=False, indent=2)

    response = requests.post(
        'https://api.anthropic.com/v1/messages',
        headers={
            'Content-Type': 'application/json',
            'x-api-key': ANTHROPIC_API_KEY,
            'anthropic-version': '2023-06-01'
        },
        json={
            'model': 'claude-sonnet-4-6',
            'max_tokens': 8000,
            'thinking': {'type': 'enabled', 'budget_tokens': 4000},
            'system': JUDGE_PROMPT,
            'messages': [{
                'role': 'user',
                'content': [
                    {
                        'type': 'image',
                        'source': {
                            'type': 'base64',
                            'media_type': media_type,
                            'data': image_b64
                        }
                    },
                    {
                        'type': 'text',
                        'text': f"Voici l'extraction du premier modèle à valider :\n\n{extraction_json}"
                    }
                ]
            }]
        },
        timeout=180
    )

    if response.status_code == 200:
        content_blocks = response.json()['content']
        for block in reversed(content_blocks):
            if block.get('type') == 'text':
                return block['text']
        raise Exception("Aucun bloc text dans la reponse Claude")

    error_msg = f"Anthropic HTTP {response.status_code}"
    try:
        error_body = response.json()
        error_detail = error_body.get('error', {}).get('message', '')
        if error_detail:
            error_msg += f" - {error_detail}"
    except Exception:
        pass
    raise Exception(error_msg)


def call_gemini_vision_judge(png_b64, extraction_claude):
    """Gemini 3 Flash Vision en mode judge (vérif extraction Claude primaire).
    Retourne le texte brut JSON."""
    if not GEMINI_API_KEY:
        raise Exception("Gemini: cle API non configuree")

    extraction_json = json.dumps(extraction_claude, ensure_ascii=False, indent=2)
    user_text = (
        f"Voici l'extraction du premier modèle à valider :\n\n{extraction_json}\n\n"
        f"Analyse l'image et corrige selon tes instructions."
    )

    response = requests.post(
        f'https://generativelanguage.googleapis.com/v1beta/models/gemini-3-flash-preview:generateContent?key={GEMINI_API_KEY}',
        headers={'Content-Type': 'application/json'},
        json={
            'system_instruction': {'parts': [{'text': JUDGE_PROMPT}]},
            'contents': [{
                'parts': [
                    {'inline_data': {'mime_type': 'image/png', 'data': png_b64}},
                    {'text': user_text}
                ]
            }],
            'generationConfig': {
                'temperature': 0.0,
                'maxOutputTokens': 8000,
                'response_mime_type': 'application/json',
            }
        },
        timeout=120
    )

    if response.status_code == 200:
        parts = response.json()['candidates'][0]['content']['parts']
        for part in parts:
            if part.get('thought'):
                continue
            if 'text' in part:
                return part['text']
        raise Exception("Gemini judge: aucune part text")
    raise Exception(f"Gemini judge HTTP {response.status_code} - {response.text[:300]}")


def needs_judge(tickets):
    """Retourne True si au moins 1 ticket nécessite la validation du judge."""
    if not tickets:
        return False
    for t in tickets:
        conf = float(t.get('confidence', 1.0) or 1.0)
        if conf < 0.80:
            return True
        ttc = float(t.get('montant_ttc', 0) or 0)
        if ttc > 200.00:
            return True
        if str(t.get('mode_paiement', '')).upper() == 'INCONNU':
            return True
        ht = float(t.get('montant_ht', 0) or 0)
        tva = float(t.get('montant_tva', 0) or 0)
        if ht > 0 and tva > 0 and abs((ht + tva) - ttc) > 0.10:
            return True
        ok, _ = check_coherence_fournisseur_type(
            t.get('fournisseur', ''), t.get('type', '')
        )
        if not ok:
            return True
    return False


# ===================================================================
# MOTEUR D'ANALYSE AVEC RETRY + FALLBACK
# ===================================================================

def clean_json_response(text, page_name=""):
    """Nettoie et parse la reponse JSON d'un LLM avec 4 niveaux de tolérance.
    Niveau 4 : json-repair pour corriger les apostrophes, guillemets mal formés, etc.
    Lève JSONDecodeError si tous les niveaux échouent."""
    # 1. Enlever les fences markdown
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text).strip()

    # 2. Extraire le premier objet JSON complet
    json_match = re.search(r'\{.*\}', text, re.DOTALL)
    if json_match:
        text = json_match.group()

    # 3. Essai direct
    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        original_error = str(e)

    # 4. Fix virgules décimales françaises dans des valeurs numériques
    fixed = re.sub(
        r'(:\s*-?\d+),(\d+)(\s*[,\}\]])',
        r'\1.\2\3',
        text
    )
    try:
        result = json.loads(fixed)
        logger.warning(f"[JSON] Fix virgules FR{' sur ' + page_name if page_name else ''}")
        return result
    except json.JSONDecodeError:
        pass

    # 5. Fix virgules de fin avant } ou ]
    fixed2 = re.sub(r',(\s*[\}\]])', r'\1', fixed)
    try:
        result = json.loads(fixed2)
        logger.warning(f"[JSON] Fix virgules de fin{' sur ' + page_name if page_name else ''}")
        return result
    except json.JSONDecodeError:
        pass

    # 6. Fix guillemets simples → doubles
    fixed3 = re.sub(r"'([^']*)':", r'"\1":', fixed2)
    try:
        result = json.loads(fixed3)
        logger.warning(f"[JSON] Fix guillemets simples{' sur ' + page_name if page_name else ''}")
        return result
    except json.JSONDecodeError:
        pass

    # 7. Troncature à la dernière } valide
    last_brace = fixed2.rfind('}')
    if last_brace > 0:
        try:
            result = json.loads(fixed2[:last_brace+1])
            logger.warning(f"[JSON] Fix troncature{' sur ' + page_name if page_name else ''}")
            return result
        except json.JSONDecodeError:
            pass

    # 8. json-repair : corrige apostrophes non échappées, guillemets mal fermés, etc.
    try:
        from json_repair import repair_json
        repaired = repair_json(text)
        result = json.loads(repaired)
        nb_tickets = len(result.get('tickets', []))
        logger.info(
            f"[JSON] json-repair réussi{' sur ' + page_name if page_name else ''} "
            f"— {nb_tickets} ticket(s) récupérés"
        )
        return result
    except Exception as e:
        logger.error(f"[JSON] json-repair échoué{' sur ' + page_name if page_name else ''}: {e}")

    # Échec total
    logger.error(f"[JSON] Impossible de parser{' : ' + page_name if page_name else ''}, début : {text[:300]}")
    logger.error(f"[JSON] Erreur originale : {original_error}")
    raise json.JSONDecodeError(original_error, text, 0)


def extract_tickets_claude_fallback(png_bytes, filename):
    """Fallback : Claude Sonnet 4.6 Vision quand Gemini échoue 2 fois.
    Utilise le même prompt vision + addon JSON strict."""
    if not ANTHROPIC_API_KEY:
        logger.warning(f"  [Fallback Claude] ANTHROPIC_API_KEY absent, impossible de fallback")
        return []

    try:
        import traceback as _traceback
        logger.info(f"  [Fallback Claude] Tentative sur {filename}")
        # compress_image_for_judge retourne (bytes, media_type, b64)
        _, media_type, image_b64 = compress_image_for_judge(png_bytes)

        prompt_fallback = VISION_PROMPT + JSON_STRICT_ADDON

        response = requests.post(
            'https://api.anthropic.com/v1/messages',
            headers={
                'Content-Type': 'application/json',
                'x-api-key': ANTHROPIC_API_KEY,
                'anthropic-version': '2023-06-01'
            },
            json={
                'model': 'claude-sonnet-4-6',
                'max_tokens': 6000,
                'system': prompt_fallback,
                'messages': [{
                    'role': 'user',
                    'content': [
                        {
                            'type': 'image',
                            'source': {
                                'type': 'base64',
                                'media_type': media_type,
                                'data': image_b64
                            }
                        },
                        {'type': 'text', 'text': 'Analyse cette image selon tes instructions.'}
                    ]
                }]
            },
            timeout=180
        )

        if response.status_code != 200:
            logger.error(f"  [Fallback Claude] HTTP {response.status_code}")
            return []

        content_blocks = response.json()['content']
        raw = next((b['text'] for b in reversed(content_blocks) if b.get('type') == 'text'), '')
        parsed = clean_json_response(raw, filename)
        tickets = parsed.get('tickets', [])
        logger.info(f"  [Fallback Claude] {len(tickets)} ticket(s) récupérés sur {filename}")
        return tickets

    except Exception as e:
        logger.error(
            f"  [Fallback Claude] Exception sur {filename}: {e}\n"
            f"{_traceback.format_exc()}"
        )
        return []

def fix_ticket_mixte_carburant(ticket, ocr_text):
    """Corrige un ticket carburant mixte (carburant + boutique) en isolant
    le montant carburant via les codes TVA TotalEnergies/ESSO (H=carburant, Q=boutique)."""
    logger.info(
        f"  [Fix mixte] Analyse ticket {ticket.get('fournisseur','?')} {ticket.get('date','?')} "
        f"ttc={ticket.get('montant_ttc')} — ocr_text[:200]={repr(ocr_text[:200])}"
    )

    if not ocr_text:
        return ticket

    has_h = bool(re.search(r'H\s*20[,.]?0*%?', ocr_text))
    has_q = bool(re.search(r'Q\s*20[,.]?0*%?', ocr_text))

    logger.info(f"  [Fix mixte] Pattern H trouvé: {has_h} / Pattern Q trouvé: {has_q}")

    if not (has_h and has_q):
        return ticket

    # Ticket mixte détecté : chercher la ligne H (carburant)
    # Format TotalEnergies : H 20,00% <HT> <TVA> <TTC>
    # Exemple : "H 20,00% 55,01 45,84 9,17" → HT=55,01  TVA=45,84  TTC=9,17 (NON)
    # Exemple réel : "H 20,00% 55,01 45,84 9,17" → groupe1=55,01 groupe2=45,84 groupe3=9,17
    match = re.search(
        r'H\s*20[,.]?0*%?\s+([\d]+[,.][\d]+)\s+([\d]+[,.][\d]+)\s+([\d]+[,.][\d]+)',
        ocr_text
    )
    if not match:
        match = re.search(
            r'H\s+20[,.]00%\s+([\d,\.]+)\s+([\d,\.]+)\s+([\d,\.]+)',
            ocr_text
        )

    if not match:
        ticket = dict(ticket)
        ticket['confidence'] = min(float(ticket.get('confidence', 1.0) or 1.0), 0.65)
        existing = ticket.get('raison_rejet', '') or ''
        note = "Ticket mixte détecté mais montant carburant non isolable - vérification manuelle requise"
        ticket['raison_rejet'] = (existing + " | " + note).strip(" |") if existing else note
        logger.warning(
            f"  [Fix mixte] {ticket.get('fournisseur','?')} {ticket.get('date','?')} : "
            f"ticket mixte carburant+boutique, ligne H non parsable"
        )
        return ticket

    # Logger les 3 valeurs capturées pour déterminer l'ordre réel (HT/TVA/TTC ou autre)
    g1 = match.group(1)
    g2 = match.group(2)
    g3 = match.group(3)
    logger.info(
        f"  [Fix mixte] Ligne H capturée — groupe1={g1} groupe2={g2} groupe3={g3} "
        f"(attendu selon TotalEnergies: HT / TVA / TTC)"
    )

    ancien_ttc = float(ticket.get('montant_ttc', 0) or 0)
    # Format TotalEnergies : H 20,00% <HT> <TVA> <TTC> → groupe3 = TTC
    nouveau_ht  = float(g1.replace(',', '.'))
    nouveau_tva = float(g2.replace(',', '.'))
    nouveau_ttc = float(g3.replace(',', '.'))

    ticket = dict(ticket)
    ticket['montant_ttc'] = nouveau_ttc
    ticket['montant_ht']  = nouveau_ht
    ticket['montant_tva'] = nouveau_tva
    # Remonter la confidence : la correction réussie valide le ticket
    ticket['confidence'] = max(0.88, float(ticket.get('confidence', 0.88) or 0.88))

    logger.warning(
        f"  [Fix mixte] {ticket.get('fournisseur','?')} {ticket.get('date','?')} : "
        f"ttc corrigé {ancien_ttc}€ → {nouveau_ttc}€ (carburant seul, boutique exclue) "
        f"→ confidence remontée à {ticket['confidence']}"
    )

    existing = ticket.get('raison_rejet', '') or ''
    note = (
        f"Ticket mixte carburant+boutique : seul le carburant "
        f"comptabilisé ({nouveau_ttc}€). Articles boutique exclus."
    )
    ticket['raison_rejet'] = (existing + " | " + note).strip(" |") if existing else note

    return ticket


def _normalize_str(s):
    """Normalise une chaîne : minuscules + suppression accents basiques."""
    replacements = str.maketrans(
        'àáâãäåèéêëìíîïòóôõöùúûüýÿçñ',
        'aaaaaaeeeeiiiiooooouuuuyyçn'
    )
    return str(s).lower().translate(replacements)

# Incohérences fournisseur/type : clés normalisées (sans accents, minuscules)
FOURNISSEUR_TYPES_INCOMPATIBLES = {
    'vinci':         {'carburant'},
    'aprr':          {'carburant'},
    'asf':           {'carburant'},
    'sanef':         {'carburant'},
    'cofiroute':     {'carburant'},
    'escota':        {'carburant'},
    'total':         {'peage'},
    'totalenergies': {'peage'},
    'esso':          {'peage'},
    'shell':         {'peage'},
    'bp':            {'peage'},
    'avia':          {'peage'},
    'ratp':          {'carburant', 'peage'},
    'sncf':          {'carburant', 'peage'},
    'free mobile':   {'carburant', 'peage', 'repas', 'hotel'},
    'orange':        {'carburant', 'peage'},
}

def check_coherence_fournisseur_type(fournisseur, type_dep):
    """Retourne (bool_coherent, raison_ou_None)."""
    if not fournisseur or not type_dep:
        return True, None
    f_norm = _normalize_str(fournisseur)
    t_norm = _normalize_str(type_dep).strip()
    for key, types_ko in FOURNISSEUR_TYPES_INCOMPATIBLES.items():
        if key in f_norm and t_norm in types_ko:
            return False, f"{fournisseur} ne vend pas de {type_dep}"
    return True, None


def extract_total_cb(ocr_text):
    """Extrait le montant total payé par carte bancaire imprimé sur le ticket.
    Retourne un float ou None si aucun pattern reconnu."""
    if not ocr_text:
        return None
    # Normaliser espaces parasites entre chiffres + virgules → points
    norm = re.sub(r'(\d)\s+(\d)', r'\1\2', ocr_text)
    norm = re.sub(r'(\d),(\d)', r'\1.\2', norm)
    patterns = [
        r'TOTAL\s*(?:CB|CARTE|PAIEMENT)\s*:?\s*([\d]+\.[\d]{2})',
        r'CARTE\s*(?:BANCAIRE|BLEUE|CB)\s*:?\s*([\d]+\.[\d]{2})',
        r'PAIEMENT\s*(?:CB|CARTE)\s*:?\s*([\d]+\.[\d]{2})',
        r'A\s*PAYER\s*:?\s*([\d]+\.[\d]{2})',
        r'MONTANT\s*(?:TTC|PAIEMENT)\s*:?\s*([\d]+\.[\d]{2})\s*(?:EUR|€)?',
        r'TOTAL\s*TTC\s*:?\s*([\d]+\.[\d]{2})',
        r'NET\s*A\s*PAYER\s*:?\s*([\d]+\.[\d]{2})',
    ]
    for p in patterns:
        matches = re.findall(p, norm, re.IGNORECASE)
        if matches:
            return max(float(m) for m in matches)
    return None


def filter_tickets_fiables(tickets, ocr_text):
    """Filtre les tickets avant calcul comptable.
    Retourne (tickets_ok, tickets_rejetes, refs_a_verifier).
    tickets_rejetes = liste de dicts ticket + 'raison_rejet'.
    refs_a_verifier = set de tuples (fournisseur, date, ttc) des tickets à confidence faible."""
    tickets_ok = []
    tickets_rejetes = []
    refs_a_verifier = set()

    # Critere 5 : OCR insuffisant -> skip les checks OCR mais ne pas rejeter
    # (pipeline vision : l'OCR DocAI est optionnel, la vision Gemini est la source primaire)
    ocr_disponible = len(ocr_text.strip()) >= 50

    acceptes = []  # pour detection doublons

    for t in tickets:
        raison = None

        # Critere 1 : montant_ttc manquant
        ttc = float(t.get('montant_ttc', 0) or 0)
        if ttc == 0:
            raison = "Montant TTC illisible ou absent"

        # Critere 2 : date manquante ou mauvais format
        elif not re.match(r'^\d{2}/\d{2}/\d{4}$', str(t.get('date', ''))):
            raison = "Date manquante ou illisible"

        # Critere 2b : dépense potentiellement personnelle → noter mais ne pas rejeter
        if raison is None:
            fournisseur_lower = str(t.get('fournisseur', '')).lower()
            is_perso = False
            for kw in ['free mobile', 'fnac', 'amazon', 'mcdo', 'mcdonald', 'auchan.fr']:
                if kw in fournisseur_lower:
                    is_perso = True
                    break
            if not is_perso and re.search(r'\bfree\b', fournisseur_lower):
                is_perso = True
            if not is_perso and re.search(r'\baction\b', fournisseur_lower):
                is_perso = True
            if not is_perso and 'carrefour' in fournisseur_lower and 'market' in fournisseur_lower:
                is_perso = True
            if is_perso:
                t = dict(t)
                t['confidence'] = min(float(t.get('confidence', 1.0) or 1.0), 0.70)
                existing = t.get('raison_rejet', '') or ''
                note = "À vérifier : dépense potentiellement personnelle"
                t['raison_rejet'] = (existing + " | " + note).strip(" |") if existing else note
                logger.info(f"  [Filtre] {t.get('fournisseur','?')} : dépense potentiellement personnelle → confidence abaissée à {t['confidence']}")

        # Fix ticket mixte carburant+boutique (avant les filtres de confiance)
        if raison is None and str(t.get('type', '')).lower() == 'carburant':
            t = fix_ticket_mixte_carburant(t, ocr_text)
            ttc = float(t.get('montant_ttc', 0) or 0)  # rafraîchir ttc après correction

        # Critere Total CB : ttc extrait ne peut pas être supérieur au montant CB imprimé
        if raison is None:
            total_cb = extract_total_cb(ocr_text)
            if total_cb is not None and ttc > 0 and ttc > total_cb + 0.05:
                t = dict(t)
                t['confidence'] = min(float(t.get('confidence', 1.0) or 1.0), 0.55)
                existing = t.get('raison_rejet', '') or ''
                note = f"TTC extrait ({ttc}) > Total CB imprimé ({total_cb})"
                t['raison_rejet'] = (existing + " | " + note).strip(" |") if existing else note
                logger.warning(
                    f"  [TotalCB] {t.get('fournisseur','?')} TTC={ttc} > CB={total_cb} "
                    f"→ confidence {t['confidence']}"
                )

        # Critere 3b : cohérence métier HT + TVA ≈ TTC (±0.02€)
        if raison is None:
            ht = float(t.get('montant_ht', 0) or 0)
            tva = float(t.get('montant_tva', 0) or 0)
            if ht > 0 and tva > 0:
                ecart = abs((ht + tva) - ttc)
                if ecart > 0.02:
                    t = dict(t)
                    t['confidence'] = min(float(t.get('confidence', 1.0) or 1.0), 0.75)
                    logger.warning(
                        f"  [Filtre] {t.get('fournisseur','?')} : HT({ht})+TVA({tva})={round(ht+tva,2)} "
                        f"≠ TTC({ttc}) (écart {ecart:.2f}€) → confidence abaissée à {t['confidence']}"
                    )

        # Critere 3c : année de la date dans [2020, année courante + 1]
        if raison is None:
            date_str = str(t.get('date', ''))
            date_match = re.match(r'^\d{2}/\d{2}/(\d{4})$', date_str)
            if date_match:
                annee = int(date_match.group(1))
                annee_max = datetime.now().year + 1
                if annee < 2020 or annee > annee_max:
                    t = dict(t)
                    t['confidence'] = min(float(t.get('confidence', 1.0) or 1.0), 0.70)
                    logger.warning(
                        f"  [Filtre] {t.get('fournisseur','?')} : année {annee} hors plage "
                        f"[2020, {annee_max}] → confidence abaissée à {t['confidence']}"
                    )

        # Critere cohérence fournisseur/type (ex: VINCI ne vend pas de carburant)
        if raison is None:
            ok, raison_incoh = check_coherence_fournisseur_type(
                t.get('fournisseur', ''), t.get('type', '')
            )
            if not ok:
                t = dict(t)
                t['confidence'] = min(float(t.get('confidence', 1.0) or 1.0), 0.50)
                existing = t.get('raison_rejet', '') or ''
                note = f"Incohérence métier : {raison_incoh}"
                t['raison_rejet'] = (existing + " | " + note).strip(" |") if existing else note
                logger.warning(
                    f"  [Cohérence] {t.get('fournisseur','?')} / {t.get('type','?')} → "
                    f"confidence {t['confidence']}"
                )

        # Critere 3 : confidence faible (après tous les ajustements)
        if raison is None:
            conf = float(t.get('confidence', 1.0) or 1.0)
            if conf <= 0.50:
                raison = f"Lecture trop incertaine (confidence {conf:.0%})"
            elif conf <= 0.70:
                t = dict(t)
                existing = t.get('raison_rejet', '') or ''
                note = f"Vérification conseillée (confidence {conf:.0%})"
                t['raison_rejet'] = (existing + " | " + note).strip(" |") if existing else note

        # Critere 4 : doublon renforcé
        if raison is None:
            fournisseur = str(t.get('fournisseur', '')).strip().lower()
            mots_f = set(fournisseur.split())
            try:
                date_t = datetime.strptime(t['date'], '%d/%m/%Y')
            except (ValueError, KeyError):
                date_t = None

            for a in acceptes:
                ttc_a = float(a.get('montant_ttc', 0) or 0)
                if abs(ttc - ttc_a) >= 0.01:
                    continue
                try:
                    date_a = datetime.strptime(a['date'], '%d/%m/%Y')
                except (ValueError, KeyError):
                    date_a = None
                if not date_t or not date_a:
                    continue

                jours = abs((date_t - date_a).days)
                fournisseur_a = str(a.get('fournisseur', '')).strip().lower()
                mots_a = set(fournisseur_a.split())

                # Doublon certain : même ttc ET même date exacte
                if jours == 0:
                    raison = "Doublon détecté"
                    break
                # Doublon probable : même ttc + date ±1 jour + mot fournisseur en commun
                if jours <= 1 and mots_f & mots_a:
                    raison = "Doublon probable (même montant, date proche, fournisseur similaire)"
                    break

        if raison:
            t_rej = dict(t)
            t_rej['raison_rejet'] = raison
            tickets_rejetes.append(t_rej)
            logger.info(f"  [Filtre] Ticket rejeté ({raison}) : {t.get('fournisseur')} {t.get('montant_ttc')}")
        else:
            conf = float(t.get('confidence', 1.0) or 1.0)
            if conf <= 0.70:
                refs_a_verifier.add((
                    str(t.get('fournisseur', '')),
                    str(t.get('date', '')),
                    float(t.get('montant_ttc', 0) or 0)
                ))
            acceptes.append(t)
            tickets_ok.append(t)

    return tickets_ok, tickets_rejetes, refs_a_verifier


def cross_validate_against_ocr(tickets, ocr_text):
    """Vérifie que chaque montant TTC est retrouvable dans le texte OCR brut.
    Si absent, abaisse la confidence à 0.70 et ajoute une raison_rejet."""
    if not ocr_text or len(ocr_text.strip()) < 50:
        return tickets

    def normalize_ocr(text):
        """Normalise le texte OCR pour la recherche de montants :
        - supprime les espaces parasites entre chiffres (ex: "1 31.55" → "131.55")
        - remplace les virgules entre chiffres par des points (ex: "131,55" → "131.55")
        """
        # Répéter 3x pour gérer les espaces multiples entre chiffres
        for _ in range(3):
            text = re.sub(r'(\d)\s+(\d)', r'\1\2', text)
        # Virgules entre chiffres → points
        text = re.sub(r'(\d),(\d)', r'\1.\2', text)
        return text

    def build_formats(amount):
        """Génère toutes les représentations normalisées d'un montant."""
        fmts = set()
        fmts.add(f"{amount:.2f}")                      # 131.55
        fmts.add(f"{amount:.1f}")                      # 131.5
        if amount == int(amount):
            fmts.add(str(int(amount)))                 # 131
        for f in list(fmts):
            fmts.add(f + "€")
            fmts.add(f + " €")
            fmts.add(f + " EUR")
        return fmts

    def find_in_ocr(amount, normalized_text, tol=0.10):
        """Cherche le montant dans le texte OCR normalisé.
        Tolérance ±0.10€ pour absorber les erreurs OCR résiduelles."""
        # Valeur exacte
        for fmt in build_formats(amount):
            if fmt in normalized_text:
                return True
        # Valeurs à ±tol
        for delta in [tol, -tol, tol/2, -tol/2]:
            candidate = round(amount + delta, 2)
            if candidate > 0:
                for fmt in build_formats(candidate):
                    if fmt in normalized_text:
                        return True
        return False

    # Normaliser le texte OCR une seule fois pour tous les tickets
    normalized_ocr = normalize_ocr(ocr_text)
    logger.info(f"  [OCR-XVal] texte normalisé[:100]={repr(normalized_ocr[:100])}")

    validated = []
    for t in tickets:
        t = dict(t)
        ttc = float(t.get('montant_ttc', 0) or 0)
        if ttc > 0 and not find_in_ocr(ttc, normalized_ocr):
            current_conf = float(t.get('confidence', 1.0) or 1.0)
            t['confidence'] = min(current_conf, 0.70)
            existing = t.get('raison_rejet', '') or ''
            note = "Montant non confirmé par OCR"
            t['raison_rejet'] = (existing + " | " + note).strip(" |") if existing else note
            logger.warning(
                f"  [OCR-XVal] {t.get('fournisseur','?')} TTC={ttc} : "
                f"montant absent du texte OCR normalisé → confidence abaissée à {t['confidence']}"
            )
        validated.append(t)
    return validated


def generate_ecritures_from_tickets(tickets, start_ref=1):
    """Moteur comptable Python pur. Equilibre garanti mathematiquement.
    L'IA extrait les donnees brutes, Python fait TOUS les calculs."""

    COMPTE_MAP = {
        'carburant': '62520000',
        'peage':     '62510000',
        'parking':   '62780000',
        'repas':     '62560000',
        'hotel':     '62560100',
        'train':     '62510000',
        'transport': '62510000',
        'fournitures': '60680000',
        'autre':     '60680000',
    }

    # 0 = non deductible, 0.80 = 80% deductible, 1.0 = 100% deductible
    TVA_RULES = {
        'carburant':   0.80,
        'peage':       1.00,
        'parking':     1.00,
        'fournitures': 1.00,
        'autre':       1.00,
        'train':       1.00,
        'transport':   1.00,
        'repas':       0,
        'hotel':       0,
    }

    ecritures = []
    alerts = []
    ref_num = start_ref

    for t in tickets:
        ref = f'T{ref_num}'
        ref_num += 1

        date = t.get('date', '')
        fournisseur = t.get('fournisseur', 'Inconnu')
        type_dep = t.get('type', 'autre').lower().strip()
        description = t.get('description', '')
        libelle = f"{fournisseur} - {description}" if description else fournisseur

        ttc = round(float(t.get('montant_ttc', 0) or 0), 2)
        tva = round(float(t.get('montant_tva', 0) or 0), 2)
        ht  = round(float(t.get('montant_ht',  0) or 0), 2)

        if ttc == 0:
            logger.warning(f"[Compta] {ref} {fournisseur}: montant_ttc=0, ticket ignore")
            ref_num -= 1
            continue

        # Recalcul si donnees partielles
        if ht == 0 and tva > 0:
            ht = round(ttc - tva, 2)
        elif tva == 0 and ht > 0:
            tva = round(ttc - ht, 2)

        # Si HT et TVA sont tous deux renseignés mais incohérents avec TTC,
        # on fait confiance au TTC (source de vérité = Total CB imprimé).
        # On recalcule HT = TTC - TVA (TVA reste fidèle au ticket imprimé).
        # Tolérance 0.05€ pour capturer les arrondis courants (ex: VINCI 19.90€).
        if ht > 0 and tva > 0 and abs((ht + tva) - ttc) > 0.005:
            ht_recalc = round(ttc - tva, 2)
            if abs(ht - ht_recalc) > 0.005:
                logger.warning(
                    f"[Compta] {ref} {fournisseur}: HT recalculé "
                    f"{ht}€ → {ht_recalc}€ (TTC={ttc} TVA={tva})"
                )
            ht = ht_recalc
            if abs((ht + tva) - ttc) > 0.02:
                # Écart résiduel trop grand → recalcul complet depuis TTC
                alerts.append(
                    f"{ref} : HT/TVA incohérents avec TTC - recalculé à partir du TTC ({ttc}€)"
                )
                taux_temp = TVA_RULES.get(type_dep, 0)
                if taux_temp > 0:
                    ht = round(ttc / 1.20, 2)
                    tva = round(ttc - ht, 2)
                else:
                    ht = ttc
                    tva = 0

        compte = COMPTE_MAP.get(type_dep, '60680000')
        taux = TVA_RULES.get(type_dep, 0)

        def ligne(cpt, debit, credit):
            return {
                'date': date, 'reference': ref, 'journal': 'FCB',
                'compte': cpt, 'libelle': libelle,
                'debit': round(debit, 2), 'credit': round(credit, 2)
            }

        if taux == 0:
            # TVA non deductible : charge = TTC
            ecritures += [ligne(compte, ttc, 0), ligne('51200000', 0, ttc)]

        elif tva == 0:
            # Pas de TVA sur le ticket : charge = TTC sans ligne TVA
            ecritures += [ligne(compte, ttc, 0), ligne('51200000', 0, ttc)]

        elif taux < 1.0:
            # Carburant : TVA partiellement deductible (80%)
            tva_ded = round(tva * taux, 2)
            tva_non_ded = round(tva * (1 - taux), 2)
            charge = round(ht + tva_non_ded, 2)
            # Ajustement centimes pour garantir debit = ttc
            if round(charge + tva_ded, 2) != ttc:
                charge = round(ttc - tva_ded, 2)
            ecritures += [ligne(compte, charge, 0), ligne('44566000', tva_ded, 0), ligne('51200000', 0, ttc)]
            alerts.append(f"{ref} : Carburant TVA 80% appliquee (defaut tourisme) - verifier si vehicule utilitaire (100%)")

        else:
            # TVA 100% deductible
            ecritures += [ligne(compte, ht, 0), ligne('44566000', tva, 0), ligne('51200000', 0, ttc)]

        # Verification equilibre — avec ajustement automatique si écart ≤ 0.02€
        lignes_ref = [e for e in ecritures if e['reference'] == ref]
        td = round(sum(e['debit'] for e in lignes_ref), 2)
        tc = round(sum(e['credit'] for e in lignes_ref), 2)
        ecart = round(tc - td, 2)
        if abs(ecart) > 0.005:
            if abs(ecart) <= 0.02:
                # Écart d'arrondi : ajuster la première ligne de charge (compte 6xxxx)
                ligne_charge = next(
                    (l for l in lignes_ref if l['compte'].startswith('6') and l['debit'] > 0),
                    None
                )
                if ligne_charge:
                    ligne_charge['debit'] = round(ligne_charge['debit'] + ecart, 2)
                    logger.info(
                        f"[Compta] {ref} {fournisseur}: équilibre ajusté "
                        f"+{ecart}€ sur ligne HT (arrondi TVA)"
                    )
                else:
                    logger.error(f"[Compta] BUG EQUILIBRE {ref}: debit={td} credit={tc}")
            else:
                logger.error(f"[Compta] BUG EQUILIBRE {ref}: debit={td} credit={tc}")

    return ecritures, ref_num - start_ref, alerts


def _majority_vote_tickets_UNUSED(runs, filename=""):  # conservé pour référence, non appelé
    """Vote champ par champ sur les tickets extraits par N runs Gemini.
    Gère les runs avec un nombre différent de tickets par exclusion du ticket extra.
    Retourne un dict JSON final avec tickets votés."""
    from collections import Counter

    if len(runs) == 1:
        return runs[0]

    def vote_field(values, numeric=False, tol=0.01):
        """Retourne (valeur_majoritaire, has_majority)."""
        if not values:
            return None, False
        if numeric:
            groups = []
            for v in values:
                merged = False
                for g in groups:
                    if abs(v - g[0]) <= tol:
                        g.append(v)
                        merged = True
                        break
                if not merged:
                    groups.append([v])
            groups.sort(key=len, reverse=True)
            if len(groups[0]) > len(values) / 2:
                return round(sum(groups[0]) / len(groups[0]), 2), True
            return values[0], False
        else:
            c = Counter(str(v) for v in values)
            best_val, best_count = c.most_common(1)[0]
            if best_count > len(values) / 2:
                return best_val, True
            return str(values[0]), False

    def ttc_present_in_others(ttc_val, other_runs, tol=0.01):
        """Vérifie si un montant TTC est présent (±tol) dans au moins un autre run."""
        for other_r in other_runs:
            for t in other_r.get('tickets', []):
                other_ttc = float(t.get('montant_ttc', 0) or 0)
                if abs(ttc_val - other_ttc) <= tol:
                    return True
        return False

    # Déterminer le nombre majoritaire de tickets
    nb_per_run = [len(r.get('tickets', [])) for r in runs]
    nb_majority = Counter(nb_per_run).most_common(1)[0][0]

    # Aligner les runs minoritaires : exclure les tickets "extra"
    aligned_runs = []
    for r in runs:
        tickets = list(r.get('tickets', []))
        if len(tickets) == nb_majority:
            aligned_runs.append(tickets)
            continue

        # Ce run a plus de tickets : identifier et exclure le(s) ticket(s) extra
        # Un ticket extra = son montant_ttc absent (±0.01€) de TOUS les autres runs
        other_runs = [other_r for other_r in runs if other_r is not r]
        filtered = [
            t for t in tickets
            if ttc_present_in_others(float(t.get('montant_ttc', 0) or 0), other_runs)
        ]
        if len(filtered) == nb_majority:
            excluded = [t for t in tickets if t not in filtered]
            for ex in excluded:
                logger.info(
                    f"  [Majority] Ticket extra exclu : {ex.get('fournisseur','?')} "
                    f"TTC={ex.get('montant_ttc')} date={ex.get('date')} "
                    f"(TTC absent dans les autres runs)"
                )
            logger.info(
                f"  [Majority] run aligné : {len(tickets)} → {len(filtered)} tickets"
            )
            aligned_runs.append(filtered)
        else:
            # Impossible d'aligner proprement : garder les nb_majority premiers
            logger.warning(
                f"  [Majority] impossible d'aligner run ({len(tickets)} tickets) → "
                f"troncature à {nb_majority}"
            )
            aligned_runs.append(tickets[:nb_majority])

    # Vote par position sur les runs alignés
    voted_tickets = []
    for idx in range(nb_majority):
        ticket_runs = [run[idx] for run in aligned_runs if idx < len(run)]
        if not ticket_runs:
            continue

        voted = {}

        # Champs texte (variations normales, pas d'impact sur confidence)
        for field in ('fournisseur', 'type', 'description', 'raison_rejet'):
            val, _ = vote_field([t.get(field, '') for t in ticket_runs], numeric=False)
            voted[field] = val

        # Date : vote strict — détermine la confidence
        voted['date'], date_majority = vote_field(
            [t.get('date', '') for t in ticket_runs], numeric=False
        )

        # Montant TTC : vote avec tolérance — détermine la confidence
        voted['montant_ttc'], ttc_majority = vote_field(
            [float(t.get('montant_ttc', 0) or 0) for t in ticket_runs], numeric=True
        )

        # Autres montants (variations normales)
        for field in ('montant_tva', 'montant_ht'):
            val, _ = vote_field(
                [float(t.get(field, 0) or 0) for t in ticket_runs], numeric=True
            )
            voted[field] = val

        # Confidence : run 1 par défaut, pénalité 0.6 UNIQUEMENT si montant_ttc diverge
        # (désaccord date seul ignoré : peut venir de tickets proches avec même montant)
        voted['confidence'] = float(ticket_runs[0].get('confidence', 1.0) or 1.0)
        if not ttc_majority:
            voted['confidence'] = min(voted['confidence'], 0.6)
            logger.warning(
                f"  [Majority] T{idx+1} ({voted.get('fournisseur','?')}) : "
                f"désaccord sur ttc → confidence abaissée à {voted['confidence']}"
            )
        elif not date_majority:
            logger.info(
                f"  [Majority] T{idx+1} ({voted.get('fournisseur','?')}) : "
                f"désaccord date uniquement (ttc unanime) → confidence inchangée"
            )

        voted_tickets.append(voted)

    # Inventaire : voter sur les champs
    inventaires = [r.get('inventaire', {}) for r in runs]
    voted_inventaire = {}
    for field in ('total_detectes', 'lisibles', 'partiels', 'illisibles'):
        vals = [inv.get(field, 0) for inv in inventaires if inv]
        if vals:
            voted_inventaire[field] = Counter(vals).most_common(1)[0][0]

    confs_globales = [float(r.get('confidence', 1.0) or 1.0) for r in runs]
    return {
        'exploitable': True,
        'inventaire': voted_inventaire,
        'tickets': voted_tickets,
        'confidence': round(sum(confs_globales) / len(confs_globales), 2)
    }


def analyze_ticket_with_retry(pdf_bytes, filename="ticket.pdf"):
    """Pipeline Vision : PDF → PNG 300dpi → Gemini Vision → [Judge Claude si besoin]."""

    # 1. Cache SHA-256 (suffixe v2 pour invalider les anciens caches texte)
    pdf_hash = hashlib.sha256(pdf_bytes).hexdigest()
    cache_file = CACHE_FOLDER / f"{pdf_hash}_v2_vision.json"
    if CACHE_ENABLED and cache_file.exists():
        try:
            cached = json.loads(cache_file.read_text(encoding='utf-8'))
            logger.info(f"[Cache] {filename} - hit vision ({pdf_hash[:8]})")
            return cached
        except Exception:
            pass

    # 2. Rendre la page en PNG 300dpi
    try:
        png_bytes, png_b64 = render_page_as_png(pdf_bytes, dpi=300)
        logger.info(f"  [Render] {filename} → PNG {len(png_bytes)//1024}KB")
    except Exception as e:
        logger.error(f"  [Render] Erreur: {e}")
        return {
            "exploitable": False,
            "raison_non_exploitable": f"Impossible de rendre le PDF: {e}",
            "tickets": []
        }

    # 3. OCR DocAI en parallèle pour cross-check (optionnel)
    ocr_text = ""
    if GOOGLE_DOCAI_PROJECT_ID and GOOGLE_DOCAI_PROCESSOR_ID:
        try:
            ocr_result = call_google_docai(pdf_bytes)
            # call_google_docai retourne une str directement (texte annoté)
            if isinstance(ocr_result, str):
                ocr_text = ocr_result
            elif isinstance(ocr_result, dict):
                ocr_text = ocr_result.get('full_text', '') or ocr_result.get('text', '')
            logger.info(f"  [OCR cross-check] {len(ocr_text)} chars")
        except Exception as e:
            logger.warning(f"  [OCR] Erreur (non-bloquant): {e}")

    # 4. Extraction Vision Gemini avec retry robuste (max 2 essais) + fallback Claude
    if not GEMINI_API_KEY and not ANTHROPIC_API_KEY:
        return {
            "exploitable": False,
            "raison_non_exploitable": "LLM indisponible",
            "tickets": []
        }

    result = None
    tickets = []
    MAX_RETRIES_PRIMARY = 2

    for attempt in range(MAX_RETRIES_PRIMARY):
        # Vérifier que le modèle primaire est disponible
        if PRIMARY_LLM == 'claude' and not ANTHROPIC_API_KEY:
            break
        if PRIMARY_LLM != 'claude' and not GEMINI_API_KEY:
            break
        try:
            # Router extraction primaire selon PRIMARY_LLM
            if PRIMARY_LLM == 'claude' and ANTHROPIC_API_KEY:
                if attempt == 0:
                    logger.info(f"[Claude Vision primary] {filename}")
                    raw = call_claude_vision_primary(png_bytes)
                else:
                    logger.warning(f"[Claude Vision primary] Retry #{attempt} sur {filename}")
                    raw = call_claude_vision_primary(png_bytes, extra_prompt=JSON_STRICT_ADDON)
            else:
                if attempt == 0:
                    logger.info(f"[Gemini Vision primary] {filename}")
                    raw = call_gemini_vision(png_b64)
                else:
                    logger.warning(f"[Gemini Vision primary] Retry #{attempt} sur {filename}")
                    raw = call_gemini_vision(png_b64, extra_prompt=JSON_STRICT_ADDON)

            result = clean_json_response(raw, filename)
            if 'tickets' not in result:
                raise ValueError("JSON sans champ 'tickets'")

            tickets = result.get('tickets', [])
            nb_vus = result.get('nb_tickets_vus', len(tickets))
            retry_tag = f" [retry #{attempt}]" if attempt > 0 else ""
            primary_label = "Claude Vision" if PRIMARY_LLM == 'claude' else "Gemini Vision"
            logger.info(f"  [{primary_label}] {len(tickets)} ticket(s){retry_tag}")

            if result.get('raisonnement'):
                logger.info(f"  [{primary_label} raisonnement] {result['raisonnement'][:200]}")
            for i, t in enumerate(tickets):
                logger.info(
                    f"  [T{i+1}] {t.get('fournisseur','?')} | "
                    f"{t.get('date','?')} | ttc={t.get('montant_ttc','?')} | "
                    f"conf={t.get('confidence','?')}"
                )

            # Détection de troncature : primaire déclare N tickets mais retourne < 50%
            if nb_vus >= 3 and len(tickets) < nb_vus * 0.5:
                logger.warning(
                    f"  [{primary_label}] Troncature probable sur {filename}: "
                    f"nb_tickets_vus={nb_vus} mais seulement {len(tickets)} extraits. "
                    f"Retry avec prompt renforcé."
                )
                result = None
                if attempt + 1 < MAX_RETRIES_PRIMARY:
                    continue
                break

            # Sous-extraction légère : log warning mais accepter
            if nb_vus > len(tickets) + 1:
                logger.warning(
                    f"  [{primary_label}] Sous-extraction légère sur {filename}: "
                    f"nb_tickets_vus={nb_vus}, extraits={len(tickets)}"
                )
            break  # Résultat cohérent — sortir de la boucle retry

        except json.JSONDecodeError as e:
            logger.error(f"[{PRIMARY_LLM} Vision] JSON invalide (essai {attempt+1}/{MAX_RETRIES_PRIMARY}): {e}")
            result = None
            if attempt + 1 < MAX_RETRIES_PRIMARY:
                continue
        except Exception as e:
            logger.error(f"[{PRIMARY_LLM} Vision] Erreur inattendue (essai {attempt+1}): {e}")
            result = None
            break  # Erreur non-JSON → pas de retry utile

    # Fallback : l'autre modèle si le primaire a échoué
    if result is None:
        if PRIMARY_LLM == 'claude':
            logger.warning(f"[Fallback] Claude primaire a échoué, bascule sur Gemini")
            if GEMINI_API_KEY:
                try:
                    raw = call_gemini_vision(png_b64, extra_prompt=JSON_STRICT_ADDON)
                    result = clean_json_response(raw, filename)
                    tickets = result.get('tickets', [])
                    logger.info(f"  [Fallback Gemini] {len(tickets)} ticket(s)")
                except Exception as e:
                    logger.error(f"  [Fallback Gemini] Echec: {e}")
                    result = None
        else:
            logger.warning(f"[Fallback] Gemini primaire a échoué, bascule sur Claude")
            tickets = extract_tickets_claude_fallback(png_bytes, filename)
            if tickets:
                result = {'tickets': tickets, 'confidence_globale': 0.75}

        if result is None or not result.get('tickets'):
            return {
                "exploitable": False,
                "raison_non_exploitable": "Extraction vision impossible (primaire + fallback échoués)",
                "tickets": []
            }
        tickets = result.get('tickets', [])

    # 5. Judge conditionnel : l'AUTRE modèle vérifie l'extraction primaire
    if tickets and needs_judge(tickets):
        judge_used = None
        raw_judge = None
        try:
            if PRIMARY_LLM == 'claude' and GEMINI_API_KEY:
                logger.info(f"  [Judge Gemini] Déclenché sur extraction Claude ({filename})")
                judge_used = 'gemini'
                raw_judge = call_gemini_vision_judge(png_b64, {
                    'tickets': tickets,
                    'confidence_globale': result.get('confidence_globale', 0.8)
                })
            elif PRIMARY_LLM != 'claude' and ANTHROPIC_API_KEY:
                logger.info(f"  [Judge Claude] Déclenché sur extraction Gemini ({filename})")
                judge_used = 'claude'
                raw_judge = call_claude_vision_judge(png_bytes, {
                    'tickets': tickets,
                    'confidence_globale': result.get('confidence_globale', 0.8)
                })

            if raw_judge:
                judge_result = clean_json_response(raw_judge)
                if 'tickets' in judge_result:
                    n_before = len(tickets)
                    tickets = judge_result['tickets']
                    modifications = judge_result.get('modifications', [])
                    logger.info(
                        f"  [Judge {judge_used}] {n_before} → {len(tickets)} tickets, "
                        f"{len(modifications)} modification(s)"
                    )
                    for mod in modifications[:5]:
                        logger.info(f"    [Judge mod] {mod}")
        except Exception as e:
            logger.warning(f"  [Judge] Echec (garde extraction primaire): {e}")

    # 6. Cross-check montants vs OCR brut
    if ocr_text:
        tickets = cross_validate_against_ocr(tickets, ocr_text)

    # 7. Sortie normalisée
    final = {
        'exploitable': True,
        'inventaire': {
            'total_detectes': len(tickets),
            'lisibles': sum(1 for t in tickets if float(t.get('confidence', 0) or 0) >= 0.80),
            'partiels': sum(1 for t in tickets if 0.50 <= float(t.get('confidence', 0) or 0) < 0.80),
            'illisibles': sum(1 for t in tickets if float(t.get('confidence', 0) or 0) < 0.50),
        },
        'tickets': tickets,
        'confidence': result.get('confidence_globale', 0.9),
        '_ocr_text': ocr_text,
    }

    # 8. Cache
    if CACHE_ENABLED:
        try:
            cache_file.write_text(
                json.dumps(final, ensure_ascii=False, indent=2),
                encoding='utf-8'
            )
            logger.info(f"[Cache] {filename} - sauvegarde vision ({pdf_hash[:8]})")
        except Exception:
            pass

    return final


# ===================================================================
# GENERATION EXCEL SAGE
# ===================================================================

def create_excel(all_ecritures, alerts=None, low_confidence_refs=None):
    """Cree le fichier Excel format Sage"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ecritures comptables"

    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    orange_fill = PatternFill(start_color='FFB347', end_color='FFB347', fill_type='solid')

    headers = ['Date', 'Reference', 'Journal', 'Compte', 'Libelle', 'Debit', 'Credit']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row = 2
    total_debit = 0
    total_credit = 0

    for e in all_ecritures:
        debit = round(float(e.get('debit', 0) or 0), 2)
        credit = round(float(e.get('credit', 0) or 0), 2)
        total_debit += debit
        total_credit += credit

        is_low_confidence = (
            low_confidence_refs and e.get('reference', '') in low_confidence_refs
        )

        values = [
            e.get('date', ''),
            e.get('reference', ''),
            e.get('journal', 'FCB'),
            e.get('compte', ''),
            e.get('libelle', ''),
            debit,
            credit
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            if is_low_confidence:
                cell.fill = orange_fill
            if col in (6, 7):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    row += 1
    equilibre = abs(total_debit - total_credit) < 0.01
    ctrl_fill = PatternFill(
        start_color='27AE60' if equilibre else 'E74C3C',
        end_color='27AE60' if equilibre else 'E74C3C',
        fill_type='solid'
    )
    ctrl_font = Font(name='Calibri', bold=True, color='FFFFFF')

    ws.cell(row=row, column=4, value='CONTROLE').font = ctrl_font
    ws.cell(row=row, column=4).fill = ctrl_fill
    status = 'OK - Equilibre' if equilibre else 'ERREUR - Desequilibre'
    ws.cell(row=row, column=5, value=status).font = ctrl_font
    ws.cell(row=row, column=5).fill = ctrl_fill
    ws.cell(row=row, column=6, value=round(total_debit, 2)).font = ctrl_font
    ws.cell(row=row, column=6).fill = ctrl_fill
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=7, value=round(total_credit, 2)).font = ctrl_font
    ws.cell(row=row, column=7).fill = ctrl_fill
    ws.cell(row=row, column=7).number_format = '#,##0.00'

    if alerts:
        row += 2
        alert_font = Font(name='Calibri', bold=True, color='E74C3C')
        ws.cell(row=row, column=1, value='ALERTES').font = alert_font
        for alert in alerts:
            row += 1
            ws.cell(row=row, column=1, value=alert)

    for col_letter, width in [('A', 14), ('B', 12), ('C', 10), ('D', 12), ('E', 45), ('F', 14), ('G', 14)]:
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ===================================================================
# RAPPORT INEXPLOITABLES
# ===================================================================

def create_inexploitable_report(inexploitable_tickets):
    """Cree un PDF listant les tickets inexploitables avec conseils de correction.
    Sépare visuellement les doublons cross-page des autres rejets."""

    CONSEILS = {
        'Scan illisible':     "Renvoyer ce ticket scanne a plat, fond blanc, sans froissures, resolution minimum 300 DPI",
        'Montant':            "Le montant TTC n'est pas visible. Verifier que le ticket n'est pas coupe ou decolore",
        'Date manquante':     "La date est absente ou illisible. Renvoyer le ticket complet",
        'Date manquante ou illisible': "La date est absente ou illisible. Renvoyer le ticket complet",
        'Lecture incertaine': "Ce ticket presente des zones illisibles. Renvoyer un scan plus net ou la version originale",
        'Doublon':            "Ce ticket semble deja present dans le lot. Verifier et renvoyer si c'est bien une depense distincte",
    }

    def get_conseil(raison):
        for key, conseil in CONSEILS.items():
            if key.lower() in raison.lower():
                return conseil
        return "Verifier ce justificatif et le renvoyer dans un prochain lot"

    # Séparer doublons cross-page et autres rejets
    doublons_cross = [t for t in inexploitable_tickets if isinstance(t, dict) and t.get('_is_doublon')]
    autres_rejets = [t for t in inexploitable_tickets if not (isinstance(t, dict) and t.get('_is_doublon'))]

    # Montants
    total_non_compta = sum(
        float(t.get('montant_ttc', 0) or 0)
        for t in autres_rejets
        if isinstance(t, dict) and float(t.get('montant_ttc', 0) or 0) > 0
    )
    total_doublons = sum(
        float(t.get('montant_ttc', 0) or 0)
        for t in doublons_cross
        if isinstance(t, dict) and float(t.get('montant_ttc', 0) or 0) > 0
    )
    nb = len(inexploitable_tickets)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    orange = (0.95, 0.5, 0.1)
    bleu = (0.2, 0.4, 0.7)

    # En-tete
    c.setFont("Helvetica-Bold", 18)
    c.setFillColor(red)
    c.drawString(50, height - 55, "Justificatifs a corriger")
    c.setFont("Helvetica", 10)
    c.setFillColor(black)
    c.drawString(50, height - 75, f"Date de traitement : {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    # Resumé chiffré
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(red)
    c.drawString(50, height - 100, f"{nb} ticket(s) non comptabilise(s)")
    if total_non_compta > 0:
        c.drawString(50, height - 116, f"Justificatifs a corriger : {total_non_compta:.2f} EUR")
    if total_doublons > 0:
        c.setFillColorRGB(*bleu)
        c.drawString(50, height - 132, f"Doublons exclus : {total_doublons:.2f} EUR ({len(doublons_cross)} ticket(s))")
    c.setFillColor(black)
    c.setFont("Helvetica", 10)
    c.drawString(50, height - 152, "Ces justificatifs n'ont pas pu etre integres. Voir les conseils ci-dessous.")

    y = height - 175

    def draw_ticket_block(ticket, is_doublon=False):
        nonlocal y
        if y < 100:
            c.showPage()
            y = height - 60

        # Ligne de separation
        c.setStrokeColorRGB(0.8, 0.8, 0.8)
        c.line(50, y + 5, width - 50, y + 5)
        y -= 5

        # Identite du ticket
        if isinstance(ticket, dict) and 'fournisseur' in ticket:
            fournisseur = ticket.get('fournisseur') or 'Inconnu'
            date_str = ticket.get('date') or 'Date manquante'
            ttc = float(ticket.get('montant_ttc', 0) or 0)
            montant_str = f" — {ttc:.2f} EUR" if ttc > 0 else ""
            identite = f"{fournisseur}  |  {date_str}{montant_str}"
            raison = ticket.get('raison_rejet', ticket.get('raison', ticket.get('raison_doublon', '')))
        else:
            identite = str(ticket.get('filename', '?'))[:60] if isinstance(ticket, dict) else str(ticket)[:60]
            raison = ticket.get('raison', '') if isinstance(ticket, dict) else ''

        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(black)
        c.drawString(52, y, identite[:80])
        y -= 16

        # Raison en orange (ou bleu pour doublons)
        c.setFont("Helvetica-Bold", 9)
        if is_doublon:
            c.setFillColorRGB(*bleu)
        else:
            c.setFillColorRGB(*orange)
        c.drawString(52, y, f"Motif : {raison[:80]}")
        y -= 14

        if is_doublon:
            c.setFont("Helvetica", 9)
            c.setFillColorRGB(0.3, 0.3, 0.3)
            c.drawString(52, y, "Verifier si ce ticket est bien un doublon avant de le renvoyer.")
        else:
            conseil = get_conseil(raison)
            c.setFont("Helvetica", 9)
            c.setFillColorRGB(0.3, 0.3, 0.3)
            if len(conseil) > 90:
                c.drawString(52, y, conseil[:90])
                y -= 12
                c.drawString(52, y, conseil[90:])
            else:
                c.drawString(52, y, conseil)
        y -= 20

    # Section 1 : Autres rejets
    if autres_rejets:
        c.setFont("Helvetica-Bold", 13)
        c.setFillColor(red)
        c.drawString(50, y, f"Justificatifs a corriger ({len(autres_rejets)})")
        y -= 20
        for ticket in autres_rejets:
            draw_ticket_block(ticket, is_doublon=False)

    # Section 2 : Doublons cross-page
    if doublons_cross:
        if y < 200:
            c.showPage()
            y = height - 60
        else:
            y -= 10

        c.setFont("Helvetica-Bold", 13)
        c.setFillColorRGB(*bleu)
        c.drawString(50, y, f"Doublons detectes et exclus ({len(doublons_cross)} tickets, {total_doublons:.2f} EUR)")
        y -= 16
        c.setFont("Helvetica", 9)
        c.setFillColorRGB(0.3, 0.3, 0.3)
        c.drawString(50, y,
            "Ces tickets ont ete identifies comme doublons d'autres tickets du lot. "
            "Verifier en cas de doute."
        )
        y -= 20
        for ticket in doublons_cross:
            draw_ticket_block(ticket, is_doublon=True)

    c.save()
    buffer.seek(0)
    return buffer.read()


# ===================================================================
# TRAITEMENT PRINCIPAL
# ===================================================================

def _normalize_fournisseur(f):
    """Normalise un nom de fournisseur pour comparaison : minuscules, sans accents,
    sans caractères spéciaux, mots-clés principaux extraits."""
    if not f:
        return ''
    s = unicodedata.normalize('NFD', str(f).lower())
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^\w\s]', ' ', s)
    stopwords = {'sarl', 'sas', 'sa', 'sasu', 'eurl', 'ets', 'cie',
                 'station', 'service', 'super', 'rel', 'relais',
                 'aire', 'autoroute', 'la', 'le', 'les', 'de', 'du', 'des'}
    mots = [m for m in s.split() if m and m not in stopwords and len(m) > 1]
    return ' '.join(sorted(mots))


def _fournisseur_similar(f1, f2, min_common_words=1):
    """Retourne True si deux noms de fournisseurs ont au moins min_common_words
    mots significatifs en commun après normalisation."""
    n1 = set(_normalize_fournisseur(f1).split())
    n2 = set(_normalize_fournisseur(f2).split())
    if not n1 or not n2:
        return False
    return len(n1 & n2) >= min_common_words


def _extract_tr_number(text):
    """Extrait le numéro Tr:XXXX des tickets péage VINCI/APRR."""
    if not text:
        return None
    m = re.search(r'Tr[:\s]*(\d{3,5})', text, re.IGNORECASE)
    return m.group(1) if m else None


def _extract_sncf_dossier(text):
    """Extrait la référence dossier SNCF (6 chars alphanumériques).
    Formats : 63UJDN, 85S394, RXZNGI, B72H2B..."""
    if not text:
        return None
    m = re.search(r'[Rr][ée]f[ée]rence\s*(?:dossier)?\s*:?\s*([A-Z0-9]{6})', text)
    if m:
        return m.group(1)
    # Format compact sans mot-clé (6 caractères alphanum majuscules seuls)
    m = re.search(r'\b([A-Z]{2}[A-Z0-9]{4}|[A-Z0-9]{2}[A-Z]{2}[A-Z0-9]{2})\b', text)
    return m.group(1) if m else None


def _extract_sens_trajet(ticket):
    """Extrait le sens du trajet péage (VINCI/APRR) : 'ENTREE→SORTIE'."""
    text = (
        str(ticket.get('description', '')) + ' ' +
        str(ticket.get('ocr_text', ''))
    ).upper()
    if not text.strip():
        return None
    # Format VINCI : "Gare X" + "Entree Y"
    m_gare = re.search(r'GARE\s*:?\s*([A-Z][A-Z\s\-]{1,20}?)(?:\s*(?:CLASSE|KM|ENTREE|$))', text)
    m_entree = re.search(r'ENTREE\s*:?\s*([A-Z][A-Z\s\-]{1,20}?)(?:\s*(?:TR|CLASSE|GARE|$))', text)
    if m_gare and m_entree:
        return f"{m_entree.group(1).strip()}>{m_gare.group(1).strip()}"
    # Format trajet explicite : "Trajet ANGERS/NANTES"
    m_trajet = re.search(r'TRAJET\s*:?\s*([A-Z]+)\s*/\s*([A-Z]+)', text)
    if m_trajet:
        return f"{m_trajet.group(1)}>{m_trajet.group(2)}"
    # Format APRR/ASF : "Entree X" + "Sortie Y"
    m_e = re.search(r'ENTREE\s*:?\s*\.?\s*([A-Z][A-Z\s\-]{1,20}?)(?:\s*(?:SORTIE|CLASSE|$))', text)
    m_s = re.search(r'SORTIE\s*:?\s*\.?\s*([A-Z][A-Z\s\-]{1,20}?)(?:\s*(?:ENTREE|CLASSE|$))', text)
    if m_e and m_s:
        return f"{m_e.group(1).strip()}>{m_s.group(1).strip()}"
    return None


def _extract_heure(ticket):
    """Extrait l'heure du ticket (HH:MM ou HH:MM:SS) comme objet datetime."""
    from datetime import datetime as _dt
    text = str(ticket.get('description', '')) + ' ' + str(ticket.get('ocr_text', ''))
    m = re.search(r'(\d{1,2}):(\d{2})(?::(\d{2}))?', text)
    if not m:
        return None
    try:
        h, mm, s = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
        if 0 <= h < 24 and 0 <= mm < 60 and 0 <= s < 60:
            return _dt(2000, 1, 1, h, mm, s)
    except ValueError:
        pass
    return None


def _extract_pompe(ticket):
    """Extrait le numéro de pompe des tickets carburant."""
    text = str(ticket.get('description', '')) + ' ' + str(ticket.get('ocr_text', ''))
    m = re.search(r'[Pp]ompe\s*[:\xb0]?\s*(\d{1,2})', text)
    return m.group(1) if m else None


def _are_distinct_tickets(t_a, t_b):
    """Retourne True si les 2 tickets sont PROUVÉS distincts malgré date+TTC+fournisseur identiques.
    Un seul discriminant positif suffit (OR logique).
    Logge la raison pour faciliter le debug."""
    # Discriminant 1 : numéros de ticket différents ET non vides
    num_a = str(t_a.get('numero_ticket', '') or '').strip()
    num_b = str(t_b.get('numero_ticket', '') or '').strip()
    if num_a and num_b and num_a != num_b:
        logger.info(f"    → PRÉSERVÉ : numéros ticket différents ({num_a} vs {num_b})")
        return True

    # Texte combiné pour les extracteurs
    text_a = str(t_a.get('description', '')) + ' ' + str(t_a.get('ocr_text', ''))
    text_b = str(t_b.get('description', '')) + ' ' + str(t_b.get('ocr_text', ''))

    # Discriminant 2 : numéro de transaction péage (Tr:XXXX) différent
    tr_a = _extract_tr_number(text_a)
    tr_b = _extract_tr_number(text_b)
    if tr_a and tr_b and tr_a != tr_b:
        logger.info(f"    → PRÉSERVÉ : Tr différents ({tr_a} vs {tr_b})")
        return True

    # Discriminant 3 : référence dossier SNCF différente
    dossier_a = _extract_sncf_dossier(text_a)
    dossier_b = _extract_sncf_dossier(text_b)
    if dossier_a and dossier_b and dossier_a != dossier_b:
        logger.info(f"    → PRÉSERVÉ : dossier SNCF différent ({dossier_a} vs {dossier_b})")
        return True

    # Discriminant 4 : sens du trajet différent (aller ≠ retour)
    sens_a = _extract_sens_trajet(t_a)
    sens_b = _extract_sens_trajet(t_b)
    if sens_a and sens_b and sens_a != sens_b:
        logger.info(f"    → PRÉSERVÉ : sens trajet différent ({sens_a} vs {sens_b})")
        return True

    # Discriminant 5 : heure différente de plus de 2 minutes
    # (<2 min = même paiement vu sur justificatif + reçu CB)
    h_a = _extract_heure(t_a)
    h_b = _extract_heure(t_b)
    if h_a and h_b:
        delta_min = abs((h_a - h_b).total_seconds() / 60)
        if delta_min > 2:
            logger.info(f"    → PRÉSERVÉ : heure différente ({delta_min:.0f} min d'écart)")
            return True

    # Discriminant 6 : numéro de pompe différent
    pompe_a = _extract_pompe(t_a)
    pompe_b = _extract_pompe(t_b)
    if pompe_a and pompe_b and pompe_a != pompe_b:
        logger.info(f"    → PRÉSERVÉ : pompe différente ({pompe_a} vs {pompe_b})")
        return True

    logger.info(f"    → DOUBLON confirmé : aucun discriminant différent")
    return False


def dedup_global_cross_page(tickets_par_page):
    """Dédup globale cross-page des tickets extraits.

    Entrée: dict {filename_page: [liste_tickets]}
    Sortie: (tickets_uniques, tickets_doublons)

    Règles :
    R1 [CERTAIN]  : même numero_ticket non vide (≥5 chars)
    R2 [CERTAIN]  : même TTC (±0.01) + même date + même fournisseur normalisé
    R3 [PROBABLE] : même TTC (±0.01) + même date + fournisseur avec ≥1 mot commun
                    + pages source différentes

    On ne déduplique PAS sur "même TTC + dates différentes" : deux péages identiques
    à quelques jours d'écart sont des dépenses distinctes et légitimes.
    """
    all_tickets = []
    for filename, tickets in tickets_par_page.items():
        for t in tickets:
            t_copy = dict(t)
            t_copy['_source_page'] = filename
            all_tickets.append(t_copy)

    logger.info(
        f"[Dedup global] Analyse de {len(all_tickets)} tickets "
        f"de {len(tickets_par_page)} pages"
    )

    tickets_uniques = []
    tickets_doublons = []

    for t in all_tickets:
        ttc = float(t.get('montant_ttc', 0) or 0)
        date = str(t.get('date', '')).strip()
        fournisseur = str(t.get('fournisseur', '')).strip()
        num_ticket = str(t.get('numero_ticket', '')).strip()
        source = t.get('_source_page', '')

        is_doublon = False
        raison_doublon = None
        doublon_de_ref = None

        for u in tickets_uniques:
            u_ttc = float(u.get('montant_ttc', 0) or 0)
            u_date = str(u.get('date', '')).strip()
            u_fournisseur = str(u.get('fournisseur', '')).strip()
            u_num = str(u.get('numero_ticket', '')).strip()
            u_source = u.get('_source_page', '')

            # R1 : numéro de ticket identique (suffisamment long pour être discriminant)
            if num_ticket and u_num and num_ticket == u_num and len(num_ticket) >= 5:
                is_doublon = True
                raison_doublon = f"Même numéro de ticket ({num_ticket})"
                doublon_de_ref = u
                break

            # Conditions communes R2/R3 : même TTC et même date
            if not (abs(ttc - u_ttc) < 0.01 and date == u_date and ttc > 0 and date):
                continue

            # R2 : fournisseur normalisé identique
            if _normalize_fournisseur(fournisseur) == _normalize_fournisseur(u_fournisseur):
                logger.info(
                    f"  [Dedup] Candidat doublon R2 : {fournisseur} {date} {ttc}€ "
                    f"({source} vs {u_source})"
                )
                if _are_distinct_tickets(t, u):
                    continue  # discriminant trouvé → tickets légitimement différents
                is_doublon = True
                raison_doublon = (
                    f"Doublon certain : {fournisseur} {date} {ttc}€ "
                    f"déjà enregistré (page {u_source})"
                )
                doublon_de_ref = u
                break

            # R3 : fournisseur similaire + pages source différentes
            if source != u_source and _fournisseur_similar(fournisseur, u_fournisseur):
                logger.info(
                    f"  [Dedup] Candidat doublon R3 : {fournisseur} ≈ {u_fournisseur} "
                    f"{date} {ttc}€ ({source} vs {u_source})"
                )
                if _are_distinct_tickets(t, u):
                    continue  # discriminant trouvé → tickets légitimement différents
                is_doublon = True
                raison_doublon = (
                    f"Doublon probable cross-page : même TTC {ttc}€ date {date}, "
                    f"fournisseur similaire ({fournisseur} ≈ {u_fournisseur}), "
                    f"pages différentes ({source} vs {u_source})"
                )
                doublon_de_ref = u
                break

        if is_doublon:
            t['raison_doublon'] = raison_doublon
            t['doublon_de'] = {
                'fournisseur': doublon_de_ref.get('fournisseur'),
                'date': doublon_de_ref.get('date'),
                'ttc': doublon_de_ref.get('montant_ttc'),
                'source': doublon_de_ref.get('_source_page'),
            } if doublon_de_ref else None
            tickets_doublons.append(t)
            logger.info(
                f"  [Dedup] DOUBLON : {fournisseur} {date} {ttc}€ "
                f"({source}) — {raison_doublon}"
            )
        else:
            tickets_uniques.append(t)

    total_doublon_montant = sum(
        float(t.get('montant_ttc', 0) or 0) for t in tickets_doublons
    )
    ratio = len(tickets_doublons) / max(len(all_tickets), 1)
    logger.info(
        f"[Dedup global] {len(tickets_uniques)} uniques, "
        f"{len(tickets_doublons)} doublons ({total_doublon_montant:.2f}€ éliminés)"
    )
    if ratio > 0.30:
        logger.warning(
            f"[Dedup global] ALERTE : {ratio:.0%} de doublons détectés — "
            f"vérifier si le PDF source contient des scans en double"
        )

    return tickets_uniques, tickets_doublons


def process_tickets(files_data):
    """Traite une liste de tickets"""
    all_ecritures = []
    exploited_pdfs = []
    inexploitable_tickets = []
    alerts = []
    low_confidence_refs = set()
    ticket_num = 1
    results_detail = []

    # Split uniquement les tres gros PDF (Claude Opus gere bien les documents multi-pages)
    split_files = []
    for file_info in files_data:
        try:
            reader = PdfReader(io.BytesIO(file_info['bytes']))
            if len(reader.pages) > 1:
                logger.info(f"Split {file_info['filename']} : {len(reader.pages)} pages")
                pages = split_pdf_pages(file_info['bytes'], file_info['filename'])
                split_files.extend(pages)
            else:
                split_files.append(file_info)
        except Exception as e:
            logger.error(f"Erreur split {file_info['filename']}: {e}")
            split_files.append(file_info)

    total_pages = len(split_files)
    logger.info(f"{'='*50}")
    logger.info(f"Traitement de {total_pages} page(s)")
    logger.info(f"{'='*50}")

    # Traitement parallèle des pages (max 4 workers)
    import time as _time
    from concurrent.futures import ThreadPoolExecutor, as_completed

    t_start = _time.monotonic()

    def process_one(idx, file_info):
        filename = file_info['filename']
        pdf_bytes = file_info['bytes']
        logger.info(f"[{idx+1}/{total_pages}] {filename}")
        result = analyze_ticket_with_retry(pdf_bytes, filename)
        return idx, filename, pdf_bytes, result

    # Lancer toutes les pages en parallèle et collecter dans l'ordre original
    ordered_results = [None] * total_pages
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {
            executor.submit(process_one, idx, file_info): idx
            for idx, file_info in enumerate(split_files)
        }
        for future in as_completed(futures):
            idx, filename, pdf_bytes, result = future.result()
            ordered_results[idx] = (filename, pdf_bytes, result)

    # ===== PASSE 1 : Filtrage qualité intra-page =====
    # Collecter tous les tickets validés par page, sans encore générer les écritures.
    tickets_par_page = {}       # {filename: [tickets_ok]}
    refs_a_verifier_par_page = {}  # {filename: set_of_refs}
    pdf_bytes_par_page = {}     # {filename: pdf_bytes}

    for filename, pdf_bytes, result in ordered_results:
        if not result.get('exploitable'):
            raison = result.get('raison_non_exploitable', 'Document inexploitable')
            inexploitable_tickets.append({'filename': filename, 'raison': raison})
            alerts.append(f"!! {filename} : {raison}")
            results_detail.append({
                'filename': filename, 'status': 'inexploitable', 'raison': raison
            })
            continue

        tickets_bruts = result.get('tickets', [])
        if not tickets_bruts:
            raison = "Aucun ticket extrait par l'IA"
            inexploitable_tickets.append({'filename': filename, 'raison': raison})
            alerts.append(f"!! {filename} : {raison}")
            results_detail.append({'filename': filename, 'status': 'inexploitable', 'raison': raison})
            continue

        ocr_text = result.get('_ocr_text', '')
        logger.info(f"  [Filtre] len(ocr_text)={len(ocr_text)} chars")
        tickets_ok, tickets_rejetes, refs_a_verifier = filter_tickets_fiables(tickets_bruts, ocr_text)
        logger.info(
            f"  [Filtre] {len(tickets_ok)} ok, {len(tickets_rejetes)} rejetes, "
            f"{len(refs_a_verifier)} à vérifier sur {len(tickets_bruts)} extraits"
        )

        # Vérification inventaire LLM
        inventaire = result.get('inventaire', {})
        total_detectes = inventaire.get('total_detectes', 0)
        nb_extraits = len(tickets_ok) + len(tickets_rejetes)
        if total_detectes > 0 and nb_extraits < total_detectes:
            manquants = total_detectes - nb_extraits
            alerts.append(
                f"!! {filename} : {manquants} ticket(s) détecté(s) par le LLM "
                f"mais non extrait(s) — vérifier le scan ({total_detectes} détectés, "
                f"{nb_extraits} traités)"
            )
            logger.warning(
                f"  [Inventaire] {total_detectes} détectés, "
                f"{nb_extraits} traités — {manquants} manquant(s)"
            )

        # Tickets rejetés intra-page → rapport inexploitables
        for t_rej in tickets_rejetes:
            raison_rej = t_rej.get('raison_rejet', 'Ticket rejete')
            label = f"{t_rej.get('fournisseur', '?')} {t_rej.get('montant_ttc', 0)}€ ({filename})"
            inexploitable_tickets.append({'filename': label, 'raison': raison_rej})
            alerts.append(f"!! {label} : {raison_rej}")

        if not tickets_ok:
            results_detail.append({
                'filename': filename, 'status': 'inexploitable',
                'raison': 'Tous les tickets rejetes par le filtre qualite'
            })
            continue

        tickets_par_page[filename] = tickets_ok
        refs_a_verifier_par_page[filename] = refs_a_verifier
        pdf_bytes_par_page[filename] = pdf_bytes

    # ===== PASSE 2 : Dédup globale cross-page =====
    tickets_uniques_par_page = {}  # {filename: [tickets_uniques]}

    if tickets_par_page and DEDUP_ENABLED:
        all_tickets_uniques, all_doublons = dedup_global_cross_page(tickets_par_page)

        # Reconstituer par page dans l'ordre original
        for t in all_tickets_uniques:
            src = t.pop('_source_page', 'unknown')
            tickets_uniques_par_page.setdefault(src, []).append(t)

        # Ajouter les doublons au rapport inexploitables
        for d in all_doublons:
            src = d.get('_source_page', 'unknown')
            label = f"{d.get('fournisseur', '?')} {d.get('montant_ttc', 0)}€ ({src})"
            inexploitable_tickets.append({
                'filename': label,
                'raison': d.get('raison_doublon', 'Doublon cross-page'),
                'montant_ttc': d.get('montant_ttc', 0),
                'fournisseur': d.get('fournisseur', ''),
                'date': d.get('date', ''),
                '_is_doublon': True,
            })
            alerts.append(f"!! [Dedup] {label} : {d.get('raison_doublon', '')}")
    else:
        # Dédup désactivée : tous les tickets passent tels quels
        for filename, tickets in tickets_par_page.items():
            tickets_uniques_par_page[filename] = [dict(t) for t in tickets]
            for t in tickets_uniques_par_page[filename]:
                t.pop('_source_page', None)

    # ===== PASSE 3 : Génération des écritures par page (ordre original) =====
    for filename, pdf_bytes, result in ordered_results:
        if filename not in tickets_uniques_par_page:
            continue  # déjà géré comme inexploitable en passe 1

        tickets_ok = tickets_uniques_par_page[filename]
        if not tickets_ok:
            results_detail.append({
                'filename': filename, 'status': 'inexploitable',
                'raison': 'Tous les tickets éliminés (rejets + doublons)'
            })
            continue

        refs_a_verifier = refs_a_verifier_par_page.get(filename, set())

        ecritures, nb_tickets, compta_alerts = generate_ecritures_from_tickets(
            tickets_ok, start_ref=ticket_num
        )
        alerts.extend(compta_alerts)
        logger.info(
            f"  [Python] {len(ecritures)} ecritures generees pour "
            f"{len(tickets_ok)} tickets ({filename})"
        )

        # Vérification d'équilibre
        par_ref = defaultdict(list)
        for e in ecritures:
            par_ref[e['reference']].append(e)
        for ref, groupe in par_ref.items():
            total_d = round(sum(e['debit'] for e in groupe), 2)
            total_c = round(sum(e['credit'] for e in groupe), 2)
            if abs(total_d - total_c) > 0.01:
                alerts.append(
                    f"{ref} ({filename}) : Desequilibre residuel "
                    f"({total_d:.2f} != {total_c:.2f})"
                )

        refs_list = ', '.join(f'T{ticket_num + i}' for i in range(nb_tickets))
        all_ecritures.extend(ecritures)

        # Surlignage orange : tickets à confidence ≤ 0.70
        for i, t_ok in enumerate(tickets_ok):
            key = (
                str(t_ok.get('fournisseur', '')),
                str(t_ok.get('date', '')),
                float(t_ok.get('montant_ttc', 0) or 0)
            )
            if key in refs_a_verifier:
                low_confidence_refs.add(f'T{ticket_num + i}')

        # TODO: reactiver le tampon S en production
        # stamped = stamp_pdf_with_s(pdf_bytes_par_page[filename])
        exploited_pdfs.append(pdf_bytes_par_page[filename])
        results_detail.append({
            'filename': filename, 'status': 'exploitable',
            'reference': refs_list, 'ecritures': ecritures
        })
        ticket_num += nb_tickets

    elapsed = _time.monotonic() - t_start
    logger.info(f"[Perf] Traitement terminé en {elapsed:.1f}s pour {total_pages} pages")

    # Generation fichiers (supprimes automatiquement apres FILE_RETENTION_MINUTES)
    output_files = {}
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if all_ecritures:
        excel_bytes = create_excel(
            all_ecritures,
            alerts if alerts else None,
            low_confidence_refs=low_confidence_refs
        )
        excel_name = f'Sage_import_{timestamp}.xlsx'
        (OUTPUT_FOLDER / excel_name).write_bytes(excel_bytes)
        output_files['excel'] = {'name': excel_name, 'path': str(OUTPUT_FOLDER / excel_name)}

    if exploited_pdfs:
        merged = merge_pdfs(exploited_pdfs)
        stamped_name = f'Tickets_exploites_S_{timestamp}.pdf'
        (OUTPUT_FOLDER / stamped_name).write_bytes(merged)
        output_files['stamped_pdf'] = {'name': stamped_name, 'path': str(OUTPUT_FOLDER / stamped_name)}

    if inexploitable_tickets:
        report = create_inexploitable_report(inexploitable_tickets)
        report_name = f'Justificatifs_inexploites_{timestamp}.pdf'
        (OUTPUT_FOLDER / report_name).write_bytes(report)
        output_files['inexploitable_pdf'] = {'name': report_name, 'path': str(OUTPUT_FOLDER / report_name)}

    total_d = round(sum(e['debit'] for e in all_ecritures), 2)
    total_c = round(sum(e['credit'] for e in all_ecritures), 2)
    logger.info(f"{'='*50}")
    logger.info(f"RESULTAT : {len(exploited_pdfs)} exploites / {len(inexploitable_tickets)} inexploitables")
    logger.info(f"TOTAUX   : D={total_d} | C={total_c} | {'OK' if abs(total_d - total_c) < 0.01 else 'ERREUR'}")
    logger.info(f"{'='*50}")

    return {
        'output_files': output_files,
        'results_detail': results_detail,
        'summary': {
            'total': total_pages,
            'exploites': len(exploited_pdfs),
            'inexploites': len(inexploitable_tickets),
            'total_debit': total_d,
            'total_credit': total_c,
            'equilibre': abs(total_d - total_c) < 0.01
        }
    }


# ===================================================================
# EMAIL
# ===================================================================

def send_email_with_attachments(to_email, subject, body, attachments):
    msg = MIMEMultipart()
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))
    for att_filename, file_bytes in attachments:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(file_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{att_filename}"')
        msg.attach(part)
    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        server.send_message(msg)


def check_emails_once():
    """Une iteration de verification des emails"""
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
    mail.select('INBOX')
    _, messages = mail.search(None, 'UNSEEN')
    for num in messages[0].split():
        if not num:
            continue
        _, msg_data = mail.fetch(num, '(RFC822)')
        msg = email.message_from_bytes(msg_data[0][1])
        sender = email.utils.parseaddr(msg['From'])[1]
        subject = msg['Subject'] or 'Sans objet'
        files_data = []
        for part in msg.walk():
            if part.get_content_type() == 'application/pdf':
                att_filename = part.get_filename() or 'document.pdf'
                pdf_bytes = part.get_payload(decode=True)
                if pdf_bytes:
                    files_data.append({'filename': att_filename, 'bytes': pdf_bytes})
        if not files_data:
            continue
        logger.info(f"[EMAIL] Mail de {sender} - {len(files_data)} PDF(s)")
        results = process_tickets(files_data)
        attachments = []
        files = results['output_files']
        for key in ['excel', 'stamped_pdf', 'inexploitable_pdf']:
            if files.get(key):
                with open(files[key]['path'], 'rb') as f:
                    attachments.append((files[key]['name'], f.read()))
        s = results['summary']
        body = f"""Bonjour,

Traitement de vos {s['total']} justificatif(s) termine.

- {s['exploites']} exploite(s)
- {s['inexploites']} inexploitable(s)
- Debit : {s['total_debit']:.2f} EUR
- Credit : {s['total_credit']:.2f} EUR
- Equilibre : {'OK' if s['equilibre'] else 'ERREUR'}

Agent Comptable IA"""
        send_email_with_attachments(sender, f"Re: {subject}", body, attachments)
        logger.info(f"[EMAIL] Reponse envoyee a {sender}")
    mail.logout()


def check_emails():
    """Boucle watchdog pour la verification des emails"""
    while True:
        try:
            check_emails_once()
        except Exception as e:
            logger.error(f"[EMAIL] Crash recupere, redemarrage dans 60s: {e}")
            time.sleep(60)
            continue
        time.sleep(CHECK_INTERVAL)


# ===================================================================
# ROUTES
# ===================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        ip = request.remote_addr

        if is_locked_out(ip):
            error = "Trop de tentatives. Reessayez dans 5 minutes."
        else:
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')

            if username == APP_USERNAME and check_password(password):
                session.permanent = True
                session['authenticated'] = True
                session['login_time'] = datetime.now().isoformat()
                session['csrf_token'] = secrets.token_hex(32)
                clear_attempts(ip)
                return redirect(url_for('index'))
            else:
                record_failed_attempt(ip)
                attempts = load_attempts()
                remaining = MAX_LOGIN_ATTEMPTS - attempts.get(ip, [0, None])[0]
                if remaining > 0:
                    error = f"Identifiants incorrects. {remaining} tentative(s) restante(s)."
                else:
                    error = "Compte bloque pour 5 minutes."

    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    return render_template('index.html', csrf_token=generate_csrf_token())


@app.route('/api/process', methods=['POST'])
@login_required
def api_process():
    # Rate limiting : max 10 appels par session par heure
    rate_key = f"{session.get('login_time', '')}_{request.remote_addr}"
    now = time.time()
    if rate_key not in PROCESS_RATE_LIMIT:
        PROCESS_RATE_LIMIT[rate_key] = []
    PROCESS_RATE_LIMIT[rate_key] = [t for t in PROCESS_RATE_LIMIT[rate_key] if now - t < 3600]
    if len(PROCESS_RATE_LIMIT[rate_key]) >= 10:
        return jsonify({'error': 'Trop de requetes, attendez avant de resoumettre'}), 429
    PROCESS_RATE_LIMIT[rate_key].append(now)

    if 'files' not in request.files:
        return jsonify({'error': 'Aucun fichier envoye'}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'Aucun fichier selectionne'}), 400

    files_data = []
    for f in files:
        if f.filename and f.filename.lower().endswith('.pdf'):
            safe_name = sanitize_filename(f.filename)
            pdf_bytes = f.read()

            # Validation : verifier que c'est bien un PDF
            if not pdf_bytes[:5] == b'%PDF-':
                continue

            files_data.append({'filename': safe_name, 'bytes': pdf_bytes})

    if not files_data:
        return jsonify({'error': 'Aucun fichier PDF valide'}), 400

    # Vérifier la limite de pages avant traitement
    total_pages_estimees = 0
    for fd in files_data:
        try:
            reader = PdfReader(io.BytesIO(fd['bytes']))
            total_pages_estimees += len(reader.pages)
        except Exception:
            total_pages_estimees += 1
    if total_pages_estimees > MAX_PAGES_PER_BATCH:
        return jsonify({
            'error': (
                f"Trop de pages ({total_pages_estimees}). "
                f"Limite : {MAX_PAGES_PER_BATCH} pages par lot. "
                f"Divise ton upload en plusieurs fichiers."
            ),
            'pages_detectees': total_pages_estimees,
            'limite': MAX_PAGES_PER_BATCH
        }), 413

    try:
        results = process_tickets(files_data)

        # Nettoyage immediat des donnees en memoire
        for fd in files_data:
            fd['bytes'] = None
        files_data.clear()

        return jsonify(results)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/download/<filename>')
@login_required
def download_file(filename):
    # Anti path-traversal
    safe_name = sanitize_filename(filename)
    filepath = OUTPUT_FOLDER / safe_name

    if not filepath.exists():
        return jsonify({'error': 'Fichier non trouve'}), 404

    # Verifier que le fichier est bien dans OUTPUT_FOLDER
    try:
        filepath.resolve().relative_to(OUTPUT_FOLDER.resolve())
    except ValueError:
        abort(403)

    @after_this_request
    def remove_file(response):
        try:
            filepath.unlink()
            logger.info(f"[ZDR] Fichier supprime apres download: {safe_name}")
        except Exception:
            pass
        return response

    return send_file(filepath, as_attachment=True, download_name=safe_name)


@app.route('/api/status')
@login_required
def api_status():
    providers = {
        'gemini': bool(GEMINI_API_KEY),
        'anthropic': bool(ANTHROPIC_API_KEY),
        'docai': bool(GOOGLE_DOCAI_PROJECT_ID and GOOGLE_DOCAI_PROCESSOR_ID),
    }

    return jsonify({
        'providers': providers,
        'active_providers': sum(1 for v in providers.values() if v),
        'file_retention_minutes': FILE_RETENTION_MINUTES,
        'max_pages_per_batch': MAX_PAGES_PER_BATCH
    })


@app.route('/api/webhook', methods=['POST'])
def webhook():
    """Endpoint webhook pour OpenClaw"""
    # Auth par token Bearer (separe du systeme de session web)
    auth_header = request.headers.get('Authorization', '')
    webhook_token = os.environ.get('WEBHOOK_TOKEN', '')

    if not webhook_token or auth_header != f'Bearer {webhook_token}':
        return jsonify({'error': 'Non autorise'}), 401

    # Accepte JSON avec base64 des PDFs
    data = request.get_json()
    if not data or 'files' not in data:
        return jsonify({'error': 'Format invalide, attendu: {"files": [{"name": "...", "data": "base64..."}]}'}), 400

    files_data = []
    for f in data['files']:
        pdf_bytes = base64.b64decode(f['data'])
        if pdf_bytes[:5] != b'%PDF-':
            continue
        files_data.append({
            'filename': sanitize_filename(f.get('name', 'document.pdf')),
            'bytes': pdf_bytes
        })

    if not files_data:
        return jsonify({'error': 'Aucun PDF valide'}), 400

    results = process_tickets(files_data)

    # Retourne le summary + les fichiers en base64
    response_data = {'summary': results['summary'], 'files': {}}
    for key, file_info in results['output_files'].items():
        with open(file_info['path'], 'rb') as fh:
            response_data['files'][key] = {
                'name': file_info['name'],
                'data': base64.b64encode(fh.read()).decode()
            }

    return jsonify(response_data)


# ===================================================================
# DEMARRAGE
# ====================================================================

if __name__ == '__main__':
    logger.info("=" * 50)
    logger.info("  AGENT COMPTABLE IA v5.0 SECURE")
    logger.info("=" * 50)

    logger.info("Securite :")
    logger.info(f"  Login         : {APP_USERNAME} / {'hash' if APP_PASSWORD_HASH else 'plain'}")
    logger.info(f"  Session       : {app.config['PERMANENT_SESSION_LIFETIME']}")
    logger.info(f"  CSRF          : actif")
    logger.info(f"  Anti-bruteforce: {MAX_LOGIN_ATTEMPTS} tentatives, lockout {LOCKOUT_DURATION}s")
    logger.info(f"  Zero Data     : fichiers supprimes apres {FILE_RETENTION_MINUTES} min")
    logger.info(f"  Headers       : CSP, X-Frame-Options, nosniff, no-cache")

    logger.info("Providers :")
    docai_ok = bool(GOOGLE_DOCAI_PROJECT_ID and GOOGLE_DOCAI_PROCESSOR_ID)
    logger.info(f"  OCR Google DocAI : {'OK' if docai_ok else 'NON'}")
    logger.info(f"  LLM Gemini       : {'OK' if GEMINI_API_KEY else 'NON'}")
    logger.info(f"  LLM Claude       : {'OK' if ANTHROPIC_API_KEY else 'NON'}")

    logger.info(f"  Webhook : {'actif sur /api/webhook' if WEBHOOK_TOKEN else 'desactive (WEBHOOK_TOKEN non defini)'}")

    if EMAIL_ADDRESS and EMAIL_PASSWORD:
        email_thread = threading.Thread(target=check_emails, daemon=True)
        email_thread.start()
        logger.info(f"Email : {EMAIL_ADDRESS}")
    else:
        logger.info("Email : non configure")

    # Thread de nettoyage automatique
    cleanup_thread = threading.Thread(target=schedule_cleanup, daemon=True)
    cleanup_thread.start()

    logger.info(f"Interface : http://localhost:5000")
    logger.info("=" * 50)

    app.run(host='0.0.0.0', port=5000, debug=False)